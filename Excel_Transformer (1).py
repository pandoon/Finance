#!/usr/bin/env python
# coding: utf-8

# In[123]:


from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
from collections import Counter
from datetime import datetime
import os
from collections import defaultdict
from openpyxl.utils import get_column_letter


# In[134]:


def is_date(value):
    """Check if a value is a valid date."""
    date_formats = ['%Y-%m-%d', '%b %d, %Y', '%Y']  # Add other formats if needed
    if isinstance(value, datetime):
        return True
    if isinstance(value, str):
        for date_format in date_formats:
            try:
                # Attempt to parse the string as a date
                datetime.strptime(value, date_format)
                return True
            except ValueError:
                continue
    return False

def find_recurring_dates(sheet):
    """Find recurring dates in columns C, D, E, and F."""
    dates_c = []
    dates_d = []
    dates_e = []
    dates_f = []

    for row in range(1, sheet.max_row + 1):
        date_c = sheet.cell(row=row, column=3).value
        date_d = sheet.cell(row=row, column=4).value
        date_e = sheet.cell(row=row, column=5).value
        date_f = sheet.cell(row=row, column=6).value

        if is_date(date_c):
            dates_c.append(date_c)
        if is_date(date_d):
            dates_d.append(date_d)
        if is_date(date_e):
            dates_e.append(date_e)
        if is_date(date_f):
            dates_f.append(date_f)

    # Find the most common date (if any)
    recurring_date_c = Counter(dates_c).most_common(1)
    recurring_date_d = Counter(dates_d).most_common(1)
    recurring_date_e = Counter(dates_e).most_common(1)
    recurring_date_f = Counter(dates_f).most_common(1)

    # Return the most common date if it occurs more than once
    date_c_header = recurring_date_c[0][0] if recurring_date_c and recurring_date_c[0][1] > 1 else None
    date_d_header = recurring_date_d[0][0] if recurring_date_d and recurring_date_d[0][1] > 1 else None
    date_e_header = recurring_date_e[0][0] if recurring_date_e and recurring_date_e[0][1] > 1 else None
    date_f_header = recurring_date_f[0][0] if recurring_date_f and recurring_date_f[0][1] > 1 else None

    return date_c_header, date_d_header, date_e_header, date_f_header

def find_tables_with_keywords(file_path, sheet_name, keywords, output_file):
    # Load the workbook and select the sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Find recurring dates in columns C, D, E, and F
    date_c_header, date_d_header, date_e_header, date_f_header = find_recurring_dates(sheet)

    # Initialize a dictionary to store table names and their associated row numbers
    tables_with_keyword_rows = {}

    # Iterate over rows in the sheet
    for row in range(1, sheet.max_row + 1):
        # Get the table name from column A
        table_name = sheet.cell(row=row, column=1).value
        # Get the cell value from column B
        cell_value = sheet.cell(row=row, column=2).value

        if cell_value:
            # Convert cell value to lowercase for case-insensitive matching
            cell_value_lower = str(cell_value).lower()

            # Check if any keyword is in the cell value
            for keyword in keywords:
                keyword_lower = keyword.lower()
                if keyword_lower in cell_value_lower:
                    # Initialize the table entry if not already present
                    if table_name not in tables_with_keyword_rows:
                        tables_with_keyword_rows[table_name] = []
                    # Add the row to the list of rows for this table
                    tables_with_keyword_rows[table_name].append(row)
                    break  # No need to check other keywords if one matches

    # Close the original workbook
    workbook.close()

    # Create a new workbook for output
    output_workbook = openpyxl.Workbook()
    output_sheet = output_workbook.active
    output_sheet.title = "Filtered Tables and Rows"

    # Write headers to the output sheet
    headers = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]
    
    # Track added headers to avoid duplication
    added_headers = set()

    # Adjust headers if recurring dates were found
    if date_c_header and date_c_header not in added_headers:
        headers[2] = date_c_header  # Column D (index 2)
        added_headers.add(date_c_header)
    if date_d_header and date_d_header not in added_headers:
        headers[3] = date_d_header  # Column E (index 3)
        added_headers.add(date_d_header)
    if date_e_header and date_e_header not in added_headers:
        headers[4] = date_e_header  # Column F (index 4)
        added_headers.add(date_e_header)
    if date_f_header and date_f_header not in added_headers:
        headers[5] = date_f_header  # Column G (index 5)
        added_headers.add(date_f_header)

    output_sheet.append(["Table Name"] + headers)

    # Reopen the workbook to read cell values for each affected row
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Track the position of inserted rows
    inserted_rows = {}

    # Write the affected rows' data to the output sheet
    for table, rows in tables_with_keyword_rows.items():
        if table not in inserted_rows:
            inserted_rows[table] = []

        for row in rows:
            # Retrieve the entire row's values from the original sheet
            row_data = [sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)]

            # Check for dates in columns C (3), D (4), E (5), and F (6)
            date_c = sheet.cell(row=row, column=3).value
            date_d = sheet.cell(row=row, column=4).value
            date_e = sheet.cell(row=row, column=5).value
            date_f = sheet.cell(row=row, column=6).value

            # Insert a new row at the top of the affected table if necessary
            if table not in inserted_rows or not inserted_rows[table]:
                # Insert an empty row before adding the first entry for the table
                output_sheet.append([""] * len(headers))
                inserted_rows[table].append(output_sheet.max_row)

            if is_date(date_c):
                output_sheet.insert_rows(inserted_rows[table][0])
                output_sheet.cell(row=inserted_rows[table][0], column=4).value = date_c
            if is_date(date_d):
                output_sheet.insert_rows(inserted_rows[table][0])
                output_sheet.cell(row=inserted_rows[table][0], column=5).value = date_d
            if is_date(date_e):
                output_sheet.insert_rows(inserted_rows[table][0])
                output_sheet.cell(row=inserted_rows[table][0], column=6).value = date_e
            if is_date(date_f):
                output_sheet.insert_rows(inserted_rows[table][0])
                output_sheet.cell(row=inserted_rows[table][0], column=7).value = date_f

            # Write the row data
            output_sheet.append([table] + row_data)

    # Close the workbook
    workbook.close()

    # Save the output workbook with a unique filename
    def get_unique_filename(base_path):
        """Generate a unique filename by appending a number if the file already exists."""
        if not os.path.exists(base_path):
            return base_path
        base, ext = os.path.splitext(base_path)
        counter = 1
        new_path = f"{base}_{counter}{ext}"
        while os.path.exists(new_path):
            counter += 1
            new_path = f"{base}_{counter}{ext}"
        return new_path

    output_file = get_unique_filename(output_file)
    output_workbook.save(output_file)


# In[164]:


file_path = r"C:\Users\thoma\Desktop\Projects\Project Oversight\Prime_NVDA.xlsx"
sheet_name = 'All_Tables'
keywords = ['Total Current Assets', 'Total Current Liabilities',
           'Total cash, cash equivalents, and short-term investments',
           'Short-term investments','Marketable securities', 'Cash and cash equivalents',
           'Total Revenue', 'Receivable',
           'Cost of Revenue','Inventory','Inventories',
           'Payable','Payables','Accounts Payable',
           'Property','Plant','Equipment',
           'Total Asset','Total Liability',
           'Gross Profit','Gross Margin',
           'Operating income', 'Net Income','stockholders’ equity',
           'Interest Expense','Tax Expense', 
           'Income before income tax', 'income tax',
           'lease payment','dividends paid','Preferred Dividend','short-term debt','Long-term debt',
           'Net income available for common shareholders','weighted average',"total stockholders' equity","total shareholders' equity",]  # Example row names
output_file = "filtered_tables.xlsx"

find_tables_with_keywords(file_path, sheet_name, keywords, output_file)


# In[ ]:




