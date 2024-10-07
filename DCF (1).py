#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
import numpy as np
import requests
import os
from datetime import datetime, timedelta
import time
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side


# In[2]:


def thick_line(sheet, start_cell, end_cell):
    """
    Applies a dotted border at the bottom of the specified cell range.

    Parameters:
    sheet (Worksheet): The worksheet to update.
    start_cell (str): The starting cell in the format 'B6'.
    end_cell (str): The ending cell in the format 'K6'.
    """
    # Convert cell addresses to row and column numbers
    start_row = int(start_cell[1:])  # Extract row number
    start_col = openpyxl.utils.column_index_from_string(start_cell[0])  # Convert column letter to index
    
    end_row = int(end_cell[1:])  # Extract row number
    end_col = openpyxl.utils.column_index_from_string(end_cell[0])  # Convert column letter to index
    
    # Create a dotted border
    dotted_border = Border(bottom=Side(style='thin'))

    # Apply the dotted border to the bottom row of the specified range
    for col in range(start_col, end_col + 1):  # Include end column
        cell = sheet.cell(row=end_row, column=col)
        cell.border = dotted_border  # Apply the dotted border


# In[3]:


def thin_line(sheet, start_cell, end_cell):
    """
    Applies a dotted border at the bottom of the specified cell range.

    Parameters:
    sheet (Worksheet): The worksheet to update.
    start_cell (str): The starting cell in the format 'B6'.
    end_cell (str): The ending cell in the format 'K6'.
    """
    # Convert cell addresses to row and column numbers
    start_row = int(start_cell[1:])  # Extract row number
    start_col = openpyxl.utils.column_index_from_string(start_cell[0])  # Convert column letter to index
    
    end_row = int(end_cell[1:])  # Extract row number
    end_col = openpyxl.utils.column_index_from_string(end_cell[0])  # Convert column letter to index
    
    # Create a dotted border
    dotted_border = Border(bottom=Side(style='dotted'))

    # Apply the dotted border to the bottom row of the specified range
    for col in range(start_col, end_col + 1):  # Include end column
        cell = sheet.cell(row=end_row, column=col)
        cell.border = dotted_border  # Apply the dotted border


# In[4]:


def set_years_in_range(sheet, start_cell, end_cell, years):
    """
    Sets the specified years in a given range of cells in an Excel sheet.

    Parameters:
    sheet (Worksheet): The worksheet to update.
    start_cell (str): The starting cell in the format 'B6'.
    end_cell (str): The ending cell in the format 'K6'.
    years (list): A list of years to fill in the specified range.
    """
    # Convert cell addresses to row and column numbers
    start_row = int(start_cell[1:])  # Extract row number
    start_col = openpyxl.utils.column_index_from_string(start_cell[0])  # Convert column letter to index
    
    end_row = int(end_cell[1:])  # Extract row number
    end_col = openpyxl.utils.column_index_from_string(end_cell[0])  # Convert column letter to index
    
    # Ensure the number of years fits within the specified range
    if (end_col - start_col + 1) < len(years):
        raise ValueError("Not enough columns in the specified range for the number of years provided.")
    
    # Set the years in the specified range
    for i, year in enumerate(years):
        sheet.cell(row=start_row, column=start_col + i).value = year


# In[5]:


def apply_thin_border(sheet, top_left_cell, bottom_right_cell):
    """
    Applies a thin border around the specified range of cells in the given sheet.

    Parameters:
    sheet (Worksheet): The worksheet to apply the border to.
    top_left_cell (str): The cell reference for the top-left corner (e.g., 'G14').
    bottom_right_cell (str): The cell reference for the bottom-right corner (e.g., 'K16').
    """
    # Convert cell references to row and column numbers
    top_left = sheet[top_left_cell]
    bottom_right = sheet[bottom_right_cell]

    # Get row and column indices
    start_row = top_left.row
    end_row = bottom_right.row
    start_col = top_left.column
    end_col = bottom_right.column

    # Define the thin border components
    thin_side = Side(style='thin')

    # Apply the border to the top row
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=start_row, column=col)
        existing_border = cell.border
        cell.border = Border(
            top=thin_side,
            left=existing_border.left,
            right=existing_border.right,
            bottom=existing_border.bottom
        )

    # Apply the border to the bottom row
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=end_row, column=col)
        existing_border = cell.border
        cell.border = Border(
            bottom=thin_side,
            left=existing_border.left,
            right=existing_border.right,
            top=existing_border.top
        )

    # Apply the border to the left column
    for row in range(start_row, end_row + 1):
        cell = sheet.cell(row=row, column=start_col)
        existing_border = cell.border
        cell.border = Border(
            left=thin_side,
            top=existing_border.top,
            right=existing_border.right,
            bottom=existing_border.bottom
        )

    # Apply the border to the right column
    for row in range(start_row, end_row + 1):
        cell = sheet.cell(row=row, column=end_col)
        existing_border = cell.border
        cell.border = Border(
            right=thin_side,
            top=existing_border.top,
            left=existing_border.left,
            bottom=existing_border.bottom
        )

    # Apply corner borders
    # Top-left corner
    top_left_cell = sheet.cell(row=start_row, column=start_col)
    top_left_cell.border = Border(top=thin_side, left=thin_side, right=top_left_cell.border.right, bottom=top_left_cell.border.bottom)

    # Top-right corner
    top_right_cell = sheet.cell(row=start_row, column=end_col)
    top_right_cell.border = Border(top=thin_side, right=thin_side, left=top_right_cell.border.left, bottom=top_right_cell.border.bottom)

    # Bottom-left corner
    bottom_left_cell = sheet.cell(row=end_row, column=start_col)
    bottom_left_cell.border = Border(bottom=thin_side, left=thin_side, top=bottom_left_cell.border.top, right=bottom_left_cell.border.right)

    # Bottom-right corner
    bottom_right_cell = sheet.cell(row=end_row, column=end_col)
    bottom_right_cell.border = Border(bottom=thin_side, right=thin_side, top=bottom_right_cell.border.top, left=bottom_right_cell.border.left)


# In[6]:


def apply_dotted_border(sheet, top_left_cell, bottom_right_cell):
    """
    Applies a dotted border around the specified range of cells in the given sheet.

    Parameters:
    sheet (Worksheet): The worksheet to apply the border to.
    top_left_cell (str): The cell reference for the top-left corner (e.g., 'G14').
    bottom_right_cell (str): The cell reference for the bottom-right corner (e.g., 'K16').
    """
    # Convert cell references to row and column numbers
    top_left = sheet[top_left_cell]
    bottom_right = sheet[bottom_right_cell]

    # Get row and column indices
    start_row = top_left.row
    end_row = bottom_right.row
    start_col = top_left.column
    end_col = bottom_right.column

    # Apply the dotted border to the outer edge
    # Top row
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=start_row, column=col)
        cell.border = Border(top=Side(style='dotted'))

    # Bottom row
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=end_row, column=col)
        cell.border = Border(bottom=Side(style='dotted'))

    # Left column
    for row in range(start_row, end_row + 1):
        cell = sheet.cell(row=row, column=start_col)
        cell.border = Border(left=Side(style='dotted'))

    # Right column
    for row in range(start_row, end_row + 1):
        cell = sheet.cell(row=row, column=end_col)
        cell.border = Border(right=Side(style='dotted'))

    # Set corners
    # Top-left corner
    sheet.cell(row=start_row, column=start_col).border = Border(top=Side(style='dotted'), left=Side(style='dotted'))

    # Top-right corner
    sheet.cell(row=start_row, column=end_col).border = Border(top=Side(style='dotted'), right=Side(style='dotted'))

    # Bottom-left corner
    sheet.cell(row=end_row, column=start_col).border = Border(bottom=Side(style='dotted'), left=Side(style='dotted'))

    # Bottom-right corner
    sheet.cell(row=end_row, column=end_col).border = Border(bottom=Side(style='dotted'), right=Side(style='dotted'))


# In[13]:


def create_dcf_file(input_path, output_path, sheet_name, company_name, years):
    """
    Reads data from a specified sheet in an input Excel file and writes a blank DCF model to a new output Excel file.

    Parameters:
    input_path (str): The path to the input Excel file.
    output_path (str): The path to the output Excel file (DCF model).
    sheet_name (str): The name of the sheet to read from the input file.
    company_name (str): The name of the company to be used as a title.
    years (list): 5 year projection time.
    """
    # Read the data from the specified sheet in the input file
    data = pd.read_excel(input_path, sheet_name=sheet_name)

    # Optional: Preview the data
    #print(data.head())

    # Create a new blank DataFrame for various sheets
    blank_df = pd.DataFrame()
    dcf_data = pd.DataFrame()
    
    years = years
    
    initial_year = years[0]
    
    four_pre = initial_year - 4
    three_pre = initial_year - 3
    two_pre = initial_year - 2
    one_pre = initial_year - 1

    # Write the blank DataFrame to a new Excel file (output file)
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        dcf_data.to_excel(writer, index=False, sheet_name='Scenarios')
        
        # Access the workbook and scenarios worksheet
        workbook = writer.book
        Scenarios = writer.sheets['Scenarios']
        
        # Define a thick bottom border
        thick_bottom_border = Border(bottom=Side(style='thick'))
        
        blank_df.to_excel(writer, index=False, sheet_name='Cover Page')
        blank_df.to_excel(writer, index=False, sheet_name='Summary')
        blank_df.to_excel(writer, index=False, sheet_name='Assumptions')
        
        blank_df.to_excel(writer, index=False, sheet_name='Model')
        
        # Access the MODEL worksheet and set titles + subtitle
        Model = writer.sheets['Model']
        Model.merge_cells('B2:O2')
        Model['B2'] = company_name
        Model.merge_cells('B3:O3')
        Model['B3'] = 'Revenue Schedule'
        merged_cellmodb2 = Model['B2']
        merged_cellmodb2.alignment = Alignment(horizontal='center', vertical='center')
        merged_cellmodb2.font = Font( size=18, bold=False)
        merged_cellmodb3 = Model['B3']
        merged_cellmodb3.alignment = Alignment(horizontal='center', vertical='center')
        merged_cellmodb3.font = Font( size=14, bold=False)
        
        # Define scenario custom header
        Model['O1'] = '="ECONOMIC: "&UPPER(CHOOSE(Scenarios!$D$6,Scenarios!C14,Scenarios!C15,Scenarios!C16))&" SCENARIO | SALES: "&UPPER(CHOOSE(Scenarios!$D$39,Scenarios!C14,Scenarios!C15,Scenarios!C16))&" SCENARIO | POLITICAL: "&UPPER(CHOOSE(Scenarios!$D$26,Scenarios!C14,Scenarios!C15,Scenarios!C16))&" SCENARIO"'
        Model['O1'].alignment = Alignment(horizontal='right')
        Model['M5'] = 'Projected'
        Model['H6'] = '=K6-3'
        Model['I6'] = '=K6-2'
        Model['J6'] = '=K6-1'
        
        Model['K6'] = '=Scenarios!G6'
        Model['L6'] = '=Scenarios!H6'
        Model['M6'] = '=Scenarios!I6'
        Model['N6'] = '=Scenarios!J6'
        Model['O6'] = '=Scenarios!K6'
        
        Model['B7'] = 'Reportable Revenue Segment A'
        Modelb7 = Model['B7']
        Modelb7.font = Font(bold=True)
        Model['C8'] = 'Annual Revenue'
        Model['E8'] = '($ MM)'
        
        Model['C9'] = 'Segment A Growth'
        Model['E9'] = '(%)'
        Model['K9'] = '=Scenarios!G44'
        Model['L9'] = '=Scenarios!H44'
        Model['M9'] = '=Scenarios!I44'
        Model['N9'] = '=Scenarios!J44'
        Model['O9'] = '=Scenarios!K44'
        
        Model['B12'] = 'Reportable Revenue Segment B'
        Modelb11 = Model['B12']
        Modelb11.font = Font(bold=True)
        Model['C13'] = 'Annual Revenue'
        Model['E13'] = '($ MM)'
        Model['C14'] = 'Segment B Growth'
        Model['E14'] = '(%)'
        Model['K14'] = '=Scenarios!G50'
        Model['L14'] = '=Scenarios!H50'
        Model['M14'] = '=Scenarios!I50'
        Model['N14'] = '=Scenarios!J50'
        Model['O14'] = '=Scenarios!K50'
        
        Model['B17'] = 'Revenue'
        Modelb15 = Model['B17']
        Modelb15.font = Font(bold=True)
        Model['C18'] = 'Segment A Revenue'
        Model['E18'] = '($ MM)'
        Model['C19'] = 'Segment B Revenue'
        Model['E19'] = '($ MM)'
        Model['C20'] = 'Net Revenue'
        Modelc18 = Model['C20']
        Modelc18.font = Font(bold=True)
        Model['E20'] = '($ MM)'
        
        Model['J20'] = '=SUM(J18:J19)'
        Model['K20'] = '=SUM(K18:J19)'
        Model['L20'] = '=SUM(L18:J19)'
        Model['M20'] = '=SUM(M18:J19)'
        Model['N20'] = '=SUM(N18:J19)'
        Model['O20'] = '=SUM(O18:J19)'
        
        
        Model['O23'] = '="ECONOMIC: "&UPPER(CHOOSE(Scenarios!$D$6,Scenarios!C14,Scenarios!C15,Scenarios!C16))&" SCENARIO | SALES: "&UPPER(CHOOSE(Scenarios!$D$39,Scenarios!C14,Scenarios!C15,Scenarios!C16))&" SCENARIO | POLITICAL: "&UPPER(CHOOSE(Scenarios!$D$26,Scenarios!C14,Scenarios!C15,Scenarios!C16))&" SCENARIO"'
        Model['O23'].alignment = Alignment(horizontal='right')
        
        Model.merge_cells('B24:O24')
        Model['B24'] = company_name
        Model.merge_cells('B25:O25')
        Model['B25'] = 'Income Statement'
        merged_cellmodb24 = Model['B24']
        merged_cellmodb24.alignment = Alignment(horizontal='center', vertical='center')
        merged_cellmodb24.font = Font( size=18, bold=False)
        merged_cellmodb25 = Model['B25']
        merged_cellmodb25.alignment = Alignment(horizontal='center', vertical='center')
        merged_cellmodb25.font = Font( size=14, bold=False)
        
        Model['B26'] = '($)'
        Model['M27'] = 'Projected'
        Model['H28'] = '=K6-3'
        Model['I28'] = '=K6-2'
        Model['J28'] = '=K6-1'
        Model['K28'] = '=Scenarios!G6'
        Model['L28'] = '=Scenarios!H6'
        Model['M28'] = '=Scenarios!I6'
        Model['N28'] = '=Scenarios!J6'
        Model['O28'] = '=Scenarios!K6'
        
        Model['C30'] = 'Revenue'
        
        Model['H30'] = data.loc[data['calendarYear'] == three_pre, 'revenue'].values[0]
        Model['I30'] = data.loc[data['calendarYear'] == two_pre, 'revenue'].values[0]
        Model['J30'] = data.loc[data['calendarYear'] == one_pre, 'revenue'].values[0]
        
        Model['C31'] = 'Cost of Revenue'
        
        Model['H31'] = data.loc[data['calendarYear'] == three_pre, 'costOfRevenue'].values[0]
        Model['I31'] = data.loc[data['calendarYear'] == two_pre, 'costOfRevenue'].values[0]
        Model['J31'] = data.loc[data['calendarYear'] == one_pre, 'costOfRevenue'].values[0]
        
        Model['K31'] = '=Assumptions!J57'
        Model['L31'] = '=Assumptions!K57'
        Model['M31'] = '=Assumptions!L57'
        Model['N31'] = '=Assumptions!M57'
        Model['O31'] = '=Assumptions!N57'
        
        Model['C32'] = 'Gross Profit'
        Modelc32 = Model['C32']
        Modelc32.font = Font(bold=True)
        
        Model['K30'] = '=K20'
        Model['L30'] = '=L20'
        Model['M30'] = '=M20'
        Model['N30'] = '=N20'
        Model['O30'] = '=O20'
        
        Model['H32'] = '=H30-H31'
        Model['I32'] = '=I30-I31'
        Model['J32'] = '=J30-J31'
        Model['K32'] = '=K30-K31'
        Model['L32'] = '=L30-L31'
        Model['M32'] = '=M30-M31'
        Model['N32'] = '=N30-N31'
        Model['O32'] = '=O30-O31'
        
        Model['C34'] = 'Research & Development Expense'
        
        Model['H34'] = data.loc[data['calendarYear'] == three_pre, 'researchAndDevelopmentExpenses'].values[0]
        Model['I34'] = data.loc[data['calendarYear'] == two_pre, 'researchAndDevelopmentExpenses'].values[0]
        Model['J34'] = data.loc[data['calendarYear'] == one_pre, 'researchAndDevelopmentExpenses'].values[0]
        
        Model['K34'] = '=Assumptions!J58'
        Model['L34'] = '=Assumptions!K58'
        Model['M34'] = '=Assumptions!L58'
        Model['N34'] = '=Assumptions!M58'
        Model['O34'] = '=Assumptions!N58'
        
        Model['C35'] = 'S G & A Expense'
        
        Model['H35'] = data.loc[data['calendarYear'] == three_pre, 'sellingGeneralAndAdministrativeExpenses'].values[0]
        Model['I35'] = data.loc[data['calendarYear'] == two_pre, 'sellingGeneralAndAdministrativeExpenses'].values[0]
        Model['J35'] = data.loc[data['calendarYear'] == one_pre, 'sellingGeneralAndAdministrativeExpenses'].values[0]
        
        Model['K35'] = '=Assumptions!J59'
        Model['L35'] = '=Assumptions!K59'
        Model['M35'] = '=Assumptions!L59'
        Model['N35'] = '=Assumptions!M59'
        Model['O35'] = '=Assumptions!N59'

        Model['C36'] = 'Operating Expenses'
        Modelc36 = Model['C36']
        Modelc36.font = Font(bold=True)
        
        Model['H36'] = '=SUM(H34:H35)'
        Model['I36'] = '=SUM(I34:I35)'
        Model['J36'] = '=SUM(J34:J35)'
        Model['K36'] = '=SUM(K34:K35)'
        Model['L36'] = '=SUM(L34:L35)'
        Model['M36'] = '=SUM(M34:M35)'
        Model['N36'] = '=SUM(N34:N35)'
        Model['O36'] = '=SUM(O34:O35)'
        
        Model['C41'] = 'EBITDA'
        Modelc41 = Model['C41']
        Modelc41.font = Font(bold=True)
        
        Model['H41'] = '=H32-H36'
        Model['I41'] = '=I32-I36'
        Model['J41'] = '=J32-J36'
        Model['K41'] = '=K32-K36'
        Model['L41'] = '=L32-L36'
        Model['M41'] = '=M32-M36'
        Model['N41'] = '=N32-N36'
        Model['O41'] = '=O32-O36'
                         
        Model['C43'] = 'Depreciation and Amortization'
        
        Model['H43'] = data.loc[data['calendarYear'] == three_pre, 'depreciationAndAmortization'].values[0]
        Model['I43'] = data.loc[data['calendarYear'] == two_pre, 'depreciationAndAmortization'].values[0]
        Model['J43'] = data.loc[data['calendarYear'] == one_pre, 'depreciationAndAmortization'].values[0]
        
        Model['C44'] = 'EBIT'
        Modelc44 = Model['C44']
        Modelc44.font = Font(bold=True)
        
        Model['K43'] = '=K145'
        Model['L43'] = '=L145'
        Model['M43'] = '=M145'
        Model['N43'] = '=N145'
        Model['O43'] = '=O145'
        
        Model['H44'] = '=H41-H43'
        Model['I44'] = '=I41-I43'
        Model['J44'] = '=J41-J43'
        Model['K44'] = '=K41-K43'
        Model['L44'] = '=L41-L43'
        Model['M44'] = '=M41-M43'
        Model['N44'] = '=N41-N43'
        Model['O44'] = '=O41-O43'
        
        Model['K46'] = '=K224'
        Model['L46'] = '=L224'
        Model['M46'] = '=M224'
        Model['N46'] = '=N224'
        Model['O46'] = '=O224'
        
        Model['H47'] = '=H44-H46'
        Model['I47'] = '=I44-I46'
        Model['J47'] = '=J44-J46'
        Model['K47'] = '=K44-K46'
        Model['L47'] = '=L44-L46'
        Model['M47'] = '=M44-M46'
        Model['N47'] = '=N44-N46'
        Model['O47'] = '=O44-O46'
        
        Model['C46'] = 'Interest Expense'
        
        Model['H46'] = data.loc[data['calendarYear'] == three_pre, 'interestExpense'].values[0]
        Model['I46'] = data.loc[data['calendarYear'] == two_pre, 'interestExpense'].values[0]
        Model['J46'] = data.loc[data['calendarYear'] == one_pre, 'interestExpense'].values[0]
        
        Model['C47'] = 'EBT'
        Modelc47 = Model['C47']
        Modelc47.font = Font(bold=True)
        
        Model['C49'] = 'Tax Expense'
        
        Model['H49'] = data.loc[data['calendarYear'] == three_pre, 'incomeTaxExpense'].values[0]
        Model['I49'] = data.loc[data['calendarYear'] == two_pre, 'incomeTaxExpense'].values[0]
        Model['J49'] = data.loc[data['calendarYear'] == one_pre, 'incomeTaxExpense'].values[0]
        
        Model['C50'] = 'Net Income'
        Modelc50 = Model['C50']
        Modelc50.font = Font(bold=True)
        
        Model['K49'] = '=K158'
        Model['L49'] = '=L158'
        Model['M49'] = '=M158'
        Model['N49'] = '=N158'
        Model['O49'] = '=O158'
        
        Model['H50'] = '=H47-H49'
        Model['I50'] = '=I47-I49'
        Model['J50'] = '=J47-J49'
        Model['K50'] = '=K47-K49'
        Model['L50'] = '=L47-L49'
        Model['M50'] = '=M47-M49'
        Model['N50'] = '=N47-N49'
        Model['O50'] = '=O47-O49'
        
        Model['B52'] = 'Margins'
        Modelb52 = Model['B52']
        Modelb52.font = Font(bold=True)
        
        Model['C53'] = 'EBITDA Margin'
        
        Model['H53'] = '=H41/H30'
        Model['I53'] = '=I41/I30'
        Model['J53'] = '=J41/J30'
        Model['K53'] = '=K41/K30'
        Model['L53'] = '=L41/L30'
        Model['M53'] = '=M41/M30'
        Model['N53'] = '=N41/N30'
        Model['O53'] = '=O41/O30'
        
        Model['C54'] = 'EBIT Margin'
        
        Model['H54'] = '=H44/H30'
        Model['I54'] = '=I44/I30'
        Model['J54'] = '=J44/J30'
        Model['K54'] = '=K44/K30'
        Model['L54'] = '=L44/L30'
        Model['M54'] = '=M44/M30'
        Model['N54'] = '=N44/N30'
        Model['O54'] = '=O44/O30'
        
        Model['C55'] = 'Return on Equity'
        
        Model['H55'] = '=H50/H95'
        Model['I55'] = '=I50/I95'
        Model['J55'] = '=J50/J95'
        Model['K55'] = '=K50/K95'
        Model['L55'] = '=L50/L95'
        Model['M55'] = '=M50/M95'
        Model['N55'] = '=N50/N95'
        Model['O55'] = '=O50/O95'
        
        Model['O58'] = '="ECONOMIC: "&UPPER(CHOOSE(Scenarios!$D$6,Scenarios!C14,Scenarios!C15,Scenarios!C16))&" SCENARIO | SALES: "&UPPER(CHOOSE(Scenarios!$D$39,Scenarios!C14,Scenarios!C15,Scenarios!C16))&" SCENARIO | POLITICAL: "&UPPER(CHOOSE(Scenarios!$D$26,Scenarios!C14,Scenarios!C15,Scenarios!C16))&" SCENARIO"'
        Model['O58'].alignment = Alignment(horizontal='right')
        
        Model.merge_cells('B59:O59')
        Model['B59'] = company_name
        Model.merge_cells('B60:O60')
        Model['B60'] = 'Balance Sheet'
        merged_cellmodb59 = Model['B59']
        merged_cellmodb59.alignment = Alignment(horizontal='center', vertical='center')
        merged_cellmodb59.font = Font( size=18, bold=False)
        merged_cellmodb60 = Model['B60']
        merged_cellmodb60.alignment = Alignment(horizontal='center', vertical='center')
        merged_cellmodb60.font = Font( size=14, bold=False)
        
        Model['B61'] = '($)'
        Model['M62'] = 'Projected'
        Model['H63'] = '=K6-3'
        Model['I63'] = '=K6-2'
        Model['J63'] = '=K6-1'
        Model['K63'] = '=Scenarios!G6'
        Model['L63'] = '=Scenarios!H6'
        Model['M63'] = '=Scenarios!I6'
        Model['N63'] = '=Scenarios!J6'
        Model['O63'] = '=Scenarios!K6'
        
        Model['B64'] = 'ASSETS'
        Modelb64 = Model['B64']
        Modelb64.font = Font(bold=True)
        Model['C65'] = 'Accounts Receivable'
        
        Model['H65'] = data.loc[data['calendarYear'] == three_pre, 'netReceivables'].values[0]
        Model['I65'] = data.loc[data['calendarYear'] == two_pre, 'netReceivables'].values[0]
        Model['J65'] = data.loc[data['calendarYear'] == one_pre, 'netReceivables'].values[0]
        
        Model['K65'] = '=K175'
        Model['L65'] = '=L175'
        Model['M65'] = '=M175'
        Model['N65'] = '=N175'
        Model['O65'] = '=O175'
        
        Model['C66'] = 'Inventory'
        
        Model['H66'] = data.loc[data['calendarYear'] == three_pre, 'inventory_balance'].values[0]
        Model['I66'] = data.loc[data['calendarYear'] == two_pre, 'inventory_balance'].values[0]
        Model['J66'] = data.loc[data['calendarYear'] == one_pre, 'inventory_balance'].values[0]
        
        Model['K66'] = '=K171'
        Model['L66'] = '=L171'
        Model['M66'] = '=M171'
        Model['N66'] = '=N171'
        Model['O66'] = '=O171'
        
        Model['C67'] = 'Cash & Cash Equivalents'
        
        Model['H67'] = data.loc[data['calendarYear'] == three_pre, 'cashAndCashEquivalents'].values[0]
        Model['I67'] = data.loc[data['calendarYear'] == two_pre, 'cashAndCashEquivalents'].values[0]
        Model['J67'] = data.loc[data['calendarYear'] == one_pre, 'cashAndCashEquivalents'].values[0]
        
        Model['K67'] = '=K127'
        Model['L67'] = '=L127'
        Model['M67'] = '=M127'
        Model['N67'] = '=N127'
        Model['O67'] = '=O127'
        
        Model['C68'] = 'Other Current Assets'
        
        Model['H68'] = data.loc[data['calendarYear'] == three_pre, 'otherCurrentAssets'].values[0] + data.loc[data['calendarYear'] == three_pre, 'shortTermInvestments'].values[0]
        Model['I68'] = data.loc[data['calendarYear'] == two_pre, 'otherCurrentAssets'].values[0] + data.loc[data['calendarYear'] == two_pre, 'shortTermInvestments'].values[0]
        Model['J68'] = data.loc[data['calendarYear'] == one_pre, 'otherCurrentAssets'].values[0] + data.loc[data['calendarYear'] == one_pre, 'shortTermInvestments'].values[0]
        
        Model['K68'] = '=K176'
        Model['L68'] = '=L176'
        Model['M68'] = '=M176'
        Model['N68'] = '=N176'
        Model['O68'] = '=O176'
        
        Model['C69'] = 'Total Current Assets'
        
        Model['H69'] = '=SUM(H65:H68)'
        Model['I69'] = '=SUM(I65:I68)'
        Model['J69'] = '=SUM(J65:J68)'
        Model['K69'] = '=SUM(K65:K68)'
        Model['L69'] = '=SUM(L65:L68)'
        Model['M69'] = '=SUM(M65:M68)'
        Model['N69'] = '=SUM(N65:N68)'
        Model['O69'] = '=SUM(O65:O68)'
        
        Modelc69 = Model['C69']
        Modelc69.font = Font(bold=True)
        
        
        
        Model['C71'] = 'Net P P & E'
        
        Model['H71'] = data.loc[data['calendarYear'] == three_pre, 'propertyPlantEquipmentNet'].values[0]
        Model['I71'] = data.loc[data['calendarYear'] == two_pre, 'propertyPlantEquipmentNet'].values[0]
        Model['J71'] = data.loc[data['calendarYear'] == one_pre, 'propertyPlantEquipmentNet'].values[0]
        
        Model['K71'] = '=J72-K116-K124'
        Model['L71'] = '=K72-L116-L124'
        Model['M71'] = '=L72-M116-M124'
        Model['N71'] = '=M72-N116-N124'
        Model['O71'] = '=N72-O116-O124'

        Model['C72'] = 'Other Non-Current Assets'
        
        Model['H72'] = data.loc[data['calendarYear'] == three_pre, 'otherNonCurrentAssets'].values[0] + data.loc[data['calendarYear'] == three_pre, 'goodwill'].values[0] + data.loc[data['calendarYear'] == three_pre, 'intangibleAssets'].values[0] + data.loc[data['calendarYear'] == three_pre, 'longTermInvestments'].values[0] + data.loc[data['calendarYear'] == three_pre, 'taxAssets'].values[0]
        Model['I72'] = data.loc[data['calendarYear'] == two_pre, 'otherNonCurrentAssets'].values[0] + data.loc[data['calendarYear'] == two_pre, 'goodwill'].values[0] + data.loc[data['calendarYear'] == two_pre, 'intangibleAssets'].values[0] + data.loc[data['calendarYear'] == two_pre, 'longTermInvestments'].values[0] + data.loc[data['calendarYear'] == two_pre, 'taxAssets'].values[0]
        Model['J72'] = data.loc[data['calendarYear'] == one_pre, 'otherNonCurrentAssets'].values[0] + data.loc[data['calendarYear'] == one_pre, 'goodwill'].values[0] + data.loc[data['calendarYear'] == one_pre, 'intangibleAssets'].values[0] + data.loc[data['calendarYear'] == one_pre, 'longTermInvestments'].values[0] + data.loc[data['calendarYear'] == one_pre, 'taxAssets'].values[0]
        
        Model['K72'] = '=J72-K115'
        Model['L72'] = '=K72-L115'
        Model['M72'] = '=L72-M115'
        Model['N72'] = '=M72-N115'
        Model['O72'] = '=N72-O115'
        
        Model['C73'] = 'Total Non-Current Assets'
        Modelc73 = Model['C73']
        Modelc73.font = Font(bold=True)
        Model['H73'] = '=SUM(H71:H72)'
        Model['I73'] = '=SUM(I71:I72)'
        Model['J73'] = '=SUM(J71:J72)'
        Model['K73'] = '=SUM(K71:K72)'
        Model['L73'] = '=SUM(L71:L72)'
        Model['M73'] = '=SUM(M71:M72)'
        Model['N73'] = '=SUM(N71:N72)'
        Model['O73'] = '=SUM(O71:O72)'
        
        Model['C75'] = 'Total Assets'
        Modelc75 = Model['C75']
        Modelc75.font = Font(bold=True)
        
        Model['H75'] = '=H73+H69'
        Model['I75'] = '=I73+I69'
        Model['J75'] = '=J73+J69'
        Model['K75'] = '=K73+K69'
        Model['L75'] = '=L73+L69'
        Model['M75'] = '=M73+M69'
        Model['N75'] = '=N73+N69'
        Model['O75'] = '=O73+O69'
        
        Model['B77'] = 'LIABILITIES AND EQUITY'
        Modelb77 = Model['B77']
        Modelb77.font = Font(bold=True)
        
        Model['C78'] = 'Accounts Payable'
        
        Model['H78'] = data.loc[data['calendarYear'] == three_pre, 'accountPayables'].values[0]
        Model['I78'] = data.loc[data['calendarYear'] == two_pre, 'accountPayables'].values[0]
        Model['J78'] = data.loc[data['calendarYear'] == one_pre, 'accountPayables'].values[0]
        
        Model['K78'] = '=K178'
        Model['L78'] = '=L178'
        Model['M78'] = '=M178'
        Model['N78'] = '=N178'
        Model['O78'] = '=O178'
        
        Model['C79'] = 'Short Term Debt'
        
        Model['H79'] = data.loc[data['calendarYear'] == three_pre, 'shortTermDebt'].values[0]
        Model['I79'] = data.loc[data['calendarYear'] == two_pre, 'shortTermDebt'].values[0]
        Model['J79'] = data.loc[data['calendarYear'] == one_pre, 'shortTermDebt'].values[0]
        Model['K79'] = '=K219'
        Model['L79'] = '=L219'
        Model['M79'] = '=M219'
        Model['N79'] = '=N219'
        Model['O79'] = '=O219'
        
        Model['C80'] = 'Other Current Liabilities'
        
        Model['H80'] = data.loc[data['calendarYear'] == three_pre, 'otherCurrentLiabilities'].values[0] + data.loc[data['calendarYear'] == three_pre, 'deferredRevenue'].values[0]
        Model['I80'] = data.loc[data['calendarYear'] == two_pre, 'otherCurrentLiabilities'].values[0] + data.loc[data['calendarYear'] == two_pre, 'deferredRevenue'].values[0]
        Model['J80'] = data.loc[data['calendarYear'] == one_pre, 'otherCurrentLiabilities'].values[0] + data.loc[data['calendarYear'] == one_pre, 'deferredRevenue'].values[0]
        
        Model['K80'] = '=K179'
        Model['L80'] = '=L179'
        Model['M80'] = '=M179'
        Model['N80'] = '=N179'
        Model['O80'] = '=O179'
        
        Model['C81'] = 'Total Current Liabilities'
        Modelc81 = Model['C81']
        Modelc81.font = Font(bold=True)
        
        Model['H81'] = '=SUM(H78:H80)'
        Model['I81'] = '=SUM(I78:I80)'
        Model['J81'] = '=SUM(J78:J80)'
        Model['K81'] = '=SUM(K78:K80)'
        Model['L81'] = '=SUM(L78:L80)'
        Model['M81'] = '=SUM(M78:M80)'
        Model['N81'] = '=SUM(N78:N80)'
        Model['O81'] = '=SUM(O78:O80)'
        
        Model['C83'] = 'Long Term Debt'
        
        Model['H83'] = data.loc[data['calendarYear'] == three_pre, 'longTermDebt'].values[0]
        Model['I83'] = data.loc[data['calendarYear'] == two_pre, 'longTermDebt'].values[0]
        Model['J83'] = data.loc[data['calendarYear'] == one_pre, 'longTermDebt'].values[0]
        Model['K83'] = '=K246'
        Model['L83'] = '=L246'
        Model['M83'] = '=M246'
        Model['N83'] = '=N246'
        Model['O83'] = '=O246'
        
        Model['C84'] = 'Other Non-Current Liabilities'
        
        Model['H84'] = data.loc[data['calendarYear'] == three_pre, 'otherNonCurrentLiabilities'].values[0] + data.loc[data['calendarYear'] == three_pre, 'deferredRevenueNonCurrent'].values[0] + data.loc[data['calendarYear'] == three_pre, 'deferredTaxLiabilitiesNonCurrent'].values[0]
        Model['I84'] = data.loc[data['calendarYear'] == two_pre, 'otherNonCurrentLiabilities'].values[0] + data.loc[data['calendarYear'] == two_pre, 'deferredRevenueNonCurrent'].values[0] + data.loc[data['calendarYear'] == two_pre, 'deferredTaxLiabilitiesNonCurrent'].values[0]
        Model['J84'] = data.loc[data['calendarYear'] == one_pre, 'otherNonCurrentLiabilities'].values[0] + data.loc[data['calendarYear'] == one_pre, 'deferredRevenueNonCurrent'].values[0] + data.loc[data['calendarYear'] == one_pre, 'deferredTaxLiabilitiesNonCurrent'].values[0]
        Model['K84'] = '=Assumptions!J48'
        Model['L84'] = '=Assumptions!K48'
        Model['M84'] = '=Assumptions!L48'
        Model['N84'] = '=Assumptions!M48'
        Model['O84'] = '=Assumptions!N48'
        
        Model['C85'] = 'Total Non-Current Liabilities'
        Modelc85 = Model['C85']
        Modelc85.font = Font(bold=True)
        
        Model['H85'] = '=SUM(H83:H84)'
        Model['I85'] = '=SUM(I83:I84)'
        Model['J85'] = '=SUM(J83:J84)'
        Model['K85'] = '=SUM(K83:K84)'
        Model['L85'] = '=SUM(L83:L84)'
        Model['M85'] = '=SUM(M83:M84)'
        Model['N85'] = '=SUM(N83:N84)'
        Model['O85'] = '=SUM(O83:O84)'
        
        Model['C87'] = 'Total Liabilities'
        Modelc87 = Model['C87']
        Modelc87.font = Font(bold=True)
        
        Model['H87'] = '=H81+H85'
        Model['I87'] = '=I81+I85'
        Model['J87'] = '=J81+J85'
        Model['K87'] = '=K81+K85'
        Model['L87'] = '=L81+L85'
        Model['M87'] = '=M81+M85'
        Model['N87'] = '=N81+N85'
        Model['O87'] = '=O81+O85'
        
        Model['C89'] = 'Preferred Shares'
        
        Model['H89'] = data.loc[data['calendarYear'] == three_pre, 'preferredStock'].values[0]
        Model['I89'] = data.loc[data['calendarYear'] == two_pre, 'preferredStock'].values[0]
        Model['J89'] = data.loc[data['calendarYear'] == one_pre, 'preferredStock'].values[0]
        
        Model['K89'] = '=K247'
        Model['L89'] = '=L247'
        Model['M89'] = '=M247'
        Model['N89'] = '=N247'
        Model['O89'] = '=O247'
        
        Model['C90'] = 'Common Shares'
        
        Model['H90'] = data.loc[data['calendarYear'] == three_pre, 'commonStock'].values[0]
        Model['I90'] = data.loc[data['calendarYear'] == two_pre, 'commonStock'].values[0]
        Model['J90'] = data.loc[data['calendarYear'] == one_pre, 'commonStock'].values[0]
        
        Model['K90'] = '=K238'
        Model['L90'] = '=L238'
        Model['M90'] = '=M238'
        Model['N90'] = '=N238'
        Model['O90'] = '=O238'
        
        Model['C91'] = 'Accumulated Other Comprehensive Income(Loss)'
        
        Model['H91'] = data.loc[data['calendarYear'] == three_pre, 'accumulatedOtherComprehensiveIncomeLoss'].values[0]
        Model['I91'] = data.loc[data['calendarYear'] == two_pre, 'accumulatedOtherComprehensiveIncomeLoss'].values[0]
        Model['J91'] = data.loc[data['calendarYear'] == one_pre, 'accumulatedOtherComprehensiveIncomeLoss'].values[0]
        
        Model['K91'] = '=K256'
        Model['L91'] = '=L256'
        Model['M91'] = '=M256'
        Model['N91'] = '=N256'
        Model['O91'] = '=O256'
        
        Model['C92'] = 'Other Stockholder Equity'
        
        Model['H92'] = data.loc[data['calendarYear'] == three_pre, 'othertotalStockholdersEquity'].values[0]
        Model['I92'] = data.loc[data['calendarYear'] == two_pre, 'othertotalStockholdersEquity'].values[0]
        Model['J92'] = data.loc[data['calendarYear'] == one_pre, 'othertotalStockholdersEquity'].values[0]
        
        Model['K92'] = '=K261'
        Model['L92'] = '=L261'
        Model['M92'] = '=M261'
        Model['N92'] = '=N261'
        Model['O92'] = '=O261'
        
        Model['C93'] = 'Retained earnings'
        
        Model['H93'] = data.loc[data['calendarYear'] == three_pre, 'retainedEarnings'].values[0]
        Model['I93'] = data.loc[data['calendarYear'] == two_pre, 'retainedEarnings'].values[0]
        Model['J93'] = data.loc[data['calendarYear'] == one_pre, 'retainedEarnings'].values[0]
        
        Model['K93'] = '=K268'
        Model['L93'] = '=L268'
        Model['M93'] = '=M268'
        Model['N93'] = '=N268'
        Model['O93'] = '=O268'
        
        Model['C94'] = 'Total Shareholder Equity'
        Modelc94 = Model['C94']
        Modelc94.font = Font(bold=True)
        
        Model['H94'] = '=SUM(H89:H93)'
        Model['I94'] = '=SUM(I89:I93)'
        Model['J94'] = '=SUM(J89:J93)'
        Model['K94'] = '=SUM(K89:K93)'
        Model['L94'] = '=SUM(L89:L93)'
        Model['M94'] = '=SUM(M89:M93)'
        Model['N94'] = '=SUM(N89:N93)'
        Model['O94'] = '=SUM(O89:O93)'
        
        Model['C96'] = 'Total Liabilities and Equity'
        Modelc96 = Model['C96']
        Modelc96.font = Font(bold=True)
        
        Model['H96'] = '=H87+H94'
        Model['I96'] = '=I87+I94'
        Model['J96'] = '=J87+J94'
        Model['K96'] = '=K87+K94'
        Model['L96'] = '=L87+L94'
        Model['M96'] = '=M87+M94'
        Model['N96'] = '=N87+N94'
        Model['O96'] = '=O87+O94'
        
        Model['O99'] = '="ECONOMIC: "&UPPER(CHOOSE(Scenarios!$D$6,Scenarios!C14,Scenarios!C15,Scenarios!C16))&" SCENARIO | SALES: "&UPPER(CHOOSE(Scenarios!$D$39,Scenarios!C14,Scenarios!C15,Scenarios!C16))&" SCENARIO | POLITICAL: "&UPPER(CHOOSE(Scenarios!$D$26,Scenarios!C14,Scenarios!C15,Scenarios!C16))&" SCENARIO"'
        Model['O99'].alignment = Alignment(horizontal='right')
        
        Model.merge_cells('B100:O100')
        Model['B100'] = company_name
        Model.merge_cells('B101:O101')
        Model['B101'] = 'Cashflow Statement'
        merged_cellmodb100 = Model['B100']
        merged_cellmodb100.alignment = Alignment(horizontal='center', vertical='center')
        merged_cellmodb100.font = Font( size=18, bold=False)
        merged_cellmodb101 = Model['B101']
        merged_cellmodb101.alignment = Alignment(horizontal='center', vertical='center')
        merged_cellmodb101.font = Font( size=14, bold=False)
        
        Model['B102'] = '($)'
        Model['M103'] = 'Projected'
        Model['H104'] = '=K6-3'
        Model['I104'] = '=K6-2'
        Model['J104'] = '=K6-1'
        Model['K104'] = '=Scenarios!G6'
        Model['L104'] = '=Scenarios!H6'
        Model['M104'] = '=Scenarios!I6'
        Model['N104'] = '=Scenarios!J6'
        Model['O104'] = '=Scenarios!K6'
        
        Model['B105'] = 'Operating Activities'
        ModelB105 = Model['B105']
        ModelB105.font = Font(bold=True)
        Model['C106'] = 'Net Income'
        Model['C107'] = 'Depreciation & Amortization'
        Model['C108'] = 'Other Non Cash Items'
        Model['C109'] = 'Change in Working Capital'
        Model['C110'] = 'Operating Cash Flow'
        ModelC110 = Model['C110']
        ModelC110.font = Font(bold=True)
        
        Model['H106'] = '=H50'
        Model['I106'] = '=I50'
        Model['J106'] = '=J50'
        Model['K106'] = '=K50'
        Model['L106'] = '=L50'
        Model['M106'] = '=M50'
        Model['N106'] = '=N50'
        Model['O106'] = '=O50'
        
        Model['H107'] = '=H43'
        Model['I107'] = '=I43'
        Model['J107'] = '=J43'
        Model['K107'] = '=K43'
        Model['L107'] = '=L43'
        Model['M107'] = '=M43'
        Model['N107'] = '=N43'
        Model['O107'] = '=O43'
        
        Model['H108'] = data.loc[data['calendarYear'] == three_pre, 'otherNonCashItems'].values[0] + data.loc[data['calendarYear'] == three_pre, 'stockBasedCompensation'].values[0] + data.loc[data['calendarYear'] == three_pre, 'deferredIncomeTax'].values[0]
        Model['I108'] = data.loc[data['calendarYear'] == two_pre, 'otherNonCashItems'].values[0] + data.loc[data['calendarYear'] == two_pre, 'stockBasedCompensation'].values[0] + data.loc[data['calendarYear'] == two_pre, 'deferredIncomeTax'].values[0]
        Model['J108'] = data.loc[data['calendarYear'] == one_pre, 'otherNonCashItems'].values[0] + data.loc[data['calendarYear'] == one_pre, 'stockBasedCompensation'].values[0] + data.loc[data['calendarYear'] == one_pre, 'deferredIncomeTax'].values[0]
        Model['K108'] = '=Assumptions!J52'
        Model['L108'] = '=Assumptions!K52'
        Model['M108'] = '=Assumptions!L52'
        Model['N108'] = '=Assumptions!M52'
        Model['O108'] = '=Assumptions!N52'
        
        
        Model['H109'] = data.loc[data['calendarYear'] == three_pre, 'changeInWorkingCapital'].values[0]
        Model['I109'] = data.loc[data['calendarYear'] == two_pre, 'changeInWorkingCapital'].values[0]
        Model['J109'] = data.loc[data['calendarYear'] == one_pre, 'changeInWorkingCapital'].values[0]
        Model['K109'] = '=K182'
        Model['L109'] = '=L182'
        Model['M109'] = '=M182'
        Model['N109'] = '=N182'
        Model['O109'] = '=O182'
        
        Model['H110'] = '=SUM(H106:H109)'
        Model['I110'] = '=SUM(I106:I109)'
        Model['J110'] = '=SUM(J106:J109)'
        Model['K110'] = '=SUM(K106:K109)'
        Model['L110'] = '=SUM(L106:L109)'
        Model['M110'] = '=SUM(M106:M109)'
        Model['N110'] = '=SUM(N106:N109)'
        Model['O110'] = '=SUM(O106:O109)'
        
        Model['B112'] = 'Investing Activities'
        ModelB112 = Model['B112']
        ModelB112.font = Font(bold=True)
        Model['C113'] = 'CAPEX'
        
        Model['H113'] = data.loc[data['calendarYear'] == three_pre, 'investmentsInPropertyPlantAndEquipment'].values[0] + data.loc[data['calendarYear'] == three_pre, 'acquisitionsNet'].values[0]
        Model['I113'] = data.loc[data['calendarYear'] == two_pre, 'investmentsInPropertyPlantAndEquipment'].values[0] + data.loc[data['calendarYear'] == two_pre, 'acquisitionsNet'].values[0]
        Model['J113'] = data.loc[data['calendarYear'] == one_pre, 'investmentsInPropertyPlantAndEquipment'].values[0] + data.loc[data['calendarYear'] == one_pre, 'acquisitionsNet'].values[0]
        Model['K113'] = '=Assumptions!J30'
        Model['L113'] = '=Assumptions!K30'
        Model['M113'] = '=Assumptions!L30'
        Model['N113'] = '=Assumptions!M30'
        Model['O113'] = '=Assumptions!N30'
        
        Model['C114'] = 'Other Investing Activities'
        
        Model['H114'] = data.loc[data['calendarYear'] == three_pre, 'otherInvestingActivites'].values[0] + data.loc[data['calendarYear'] == three_pre, 'purchasesOfInvestments'].values[0] + data.loc[data['calendarYear'] == three_pre, 'salesMaturitiesOfInvestments'].values[0]
        Model['I114'] = data.loc[data['calendarYear'] == two_pre, 'otherInvestingActivites'].values[0] + data.loc[data['calendarYear'] == two_pre, 'purchasesOfInvestments'].values[0] + data.loc[data['calendarYear'] == two_pre, 'salesMaturitiesOfInvestments'].values[0]
        Model['J114'] = data.loc[data['calendarYear'] == one_pre, 'otherInvestingActivites'].values[0] + data.loc[data['calendarYear'] == one_pre, 'purchasesOfInvestments'].values[0] + data.loc[data['calendarYear'] == one_pre, 'salesMaturitiesOfInvestments'].values[0]
        Model['K114'] = '=Assumptions!J31'
        Model['L114'] = '=Assumptions!K31'
        Model['M114'] = '=Assumptions!L31'
        Model['N114'] = '=Assumptions!M31'
        Model['O114'] = '=Assumptions!N31'
        
        Model['C115'] = 'Investing Cash Flow'
        ModelC115 = Model['C115']
        ModelC115.font = Font(bold=True)
        
        Model['H115'] = '=SUM(H113:H114)'
        Model['I115'] = '=SUM(I113:I114)'
        Model['J115'] = '=SUM(J113:J114)'
        Model['K115'] = '=SUM(K113:K114)'
        Model['L115'] = '=SUM(L113:L114)'
        Model['M115'] = '=SUM(M113:M114)'
        Model['N115'] = '=SUM(N113:N114)'
        Model['O115'] = '=SUM(O113:O114)'
        
        Model['B117'] = 'Financing Activities'
        ModelB117 = Model['B117']
        ModelB117.font = Font(bold=True)
        Model['C118'] = 'Short Term Debt Issuance (Payment)'
        
        Model['H118'] = data.loc[data['calendarYear'] == three_pre, 'debtRepayment'].values[0]
        Model['I118'] = data.loc[data['calendarYear'] == two_pre, 'debtRepayment'].values[0]
        Model['J118'] = data.loc[data['calendarYear'] == one_pre, 'debtRepayment'].values[0]
        Model['K118'] = '=K218'
        Model['L118'] = '=L218'
        Model['M118'] = '=M218'
        Model['N118'] = '=N218'
        Model['O118'] = '=O218'
        
        Model['C119'] = 'Long Term Debt Issuance (Payment)'
        
        Model['K119'] = '=K203'
        Model['L119'] = '=L203'
        Model['M119'] = '=M203'
        Model['N119'] = '=N203'
        Model['O119'] = '=O203'
        
        Model['C120'] = 'Repurchases of Common Stock'
        
        Model['H120'] = data.loc[data['calendarYear'] == three_pre, 'commonStockRepurchased'].values[0]
        Model['I120'] = data.loc[data['calendarYear'] == two_pre, 'commonStockRepurchased'].values[0]
        Model['J120'] = data.loc[data['calendarYear'] == one_pre, 'commonStockRepurchased'].values[0]
        Model['K120'] = '=K237'
        Model['L120'] = '=L237'
        Model['M120'] = '=M237'
        Model['N120'] = '=N237'
        Model['O120'] = '=O237'
        
        Model['C121'] = 'Dividends Paid'
        
        Model['H121'] = data.loc[data['calendarYear'] == three_pre, 'dividendsPaid'].values[0]
        Model['I121'] = data.loc[data['calendarYear'] == two_pre, 'dividendsPaid'].values[0]
        Model['J121'] = data.loc[data['calendarYear'] == one_pre, 'dividendsPaid'].values[0]
        Model['K121'] = '=K214'
        Model['L121'] = '=L214'
        Model['M121'] = '=M214'
        Model['N121'] = '=N214'
        Model['O121'] = '=O214'
        
        Model['C122'] = 'Other Financing Activities'
        
        Model['H122'] = data.loc[data['calendarYear'] == three_pre, 'otherFinancingActivites'].values[0] + data.loc[data['calendarYear'] == three_pre, 'commonStockIssued'].values[0]
        Model['I122'] = data.loc[data['calendarYear'] == two_pre, 'otherFinancingActivites'].values[0] + data.loc[data['calendarYear'] == three_pre, 'commonStockIssued'].values[0]
        Model['J122'] = data.loc[data['calendarYear'] == one_pre, 'otherFinancingActivites'].values[0] + data.loc[data['calendarYear'] == three_pre, 'commonStockIssued'].values[0]
        Model['K122'] = '=Assumptions!J54'
        Model['L122'] = '=Assumptions!K54'
        Model['M122'] = '=Assumptions!L54'
        Model['N122'] = '=Assumptions!M54'
        Model['O122'] = '=Assumptions!N54'
        
        Model['C123'] = 'Financing Cash Flow'
        ModelC123 = Model['C123']
        ModelC123.font = Font(bold=True)
        
        Model['H123'] = '=SUM(H118:H122)'
        Model['I123'] = '=SUM(I118:I122)'
        Model['J123'] = '=SUM(J118:J122)'
        Model['K123'] = '=SUM(K118:K122)'
        Model['L123'] = '=SUM(L118:L122)'
        Model['M123'] = '=SUM(M118:M122)'
        Model['N123'] = '=SUM(N118:N122)'
        Model['O123'] = '=SUM(O118:O122)'
        
        Model['B125'] = 'Change in Cash Position'
        
        Model['H125'] = '=H110+H123+H115'
        Model['I125'] = '=I110+I123+I115'
        Model['J125'] = '=J110+J123+J115'
        Model['K125'] = '=K110+K123+K115'
        Model['L125'] = '=L110+L123+L115'
        Model['M125'] = '=M110+M123+M115'
        Model['N125'] = '=N110+N123+N115'
        Model['O125'] = '=O110+O123+O115'
        
        Model['H127'] = '=SUM(H125:H126)'
        Model['I127'] = '=SUM(I125:I126)'
        Model['J127'] = '=SUM(J125:J126)'
        Model['K127'] = '=SUM(K125:K126)'
        Model['L127'] = '=SUM(L125:L126)'
        Model['M127'] = '=SUM(M125:M126)'
        Model['N127'] = '=SUM(N125:N126)'
        Model['O127'] = '=SUM(O125:O126)'
        
        Model['H126'] = data.loc[data['calendarYear'] == four_pre, 'cashAndCashEquivalents'].values[0]
        Model['I126'] = '=H127'
        Model['J126'] = '=I127'
        Model['K126'] = '=J127'
        Model['L126'] = '=K127'
        Model['M126'] = '=L127'
        Model['N126'] = '=M127'
        Model['O126'] = '=N127'
        
        Model['B126'] = 'Starting Cash Position'
        Model['B127'] = 'Ending Cash Position'
        
        Model['O130'] = '="ECONOMIC: "&UPPER(CHOOSE(Scenarios!$D$6,Scenarios!C14,Scenarios!C15,Scenarios!C16))&" SCENARIO | SALES: "&UPPER(CHOOSE(Scenarios!$D$39,Scenarios!C14,Scenarios!C15,Scenarios!C16))&" SCENARIO | POLITICAL: "&UPPER(CHOOSE(Scenarios!$D$26,Scenarios!C14,Scenarios!C15,Scenarios!C16))&" SCENARIO"'
        Model['O130'].alignment = Alignment(horizontal='right')
        
        Model.merge_cells('B131:O131')
        Model['B131'] = company_name
        Model.merge_cells('B132:O132')
        Model['B132'] = 'Depreciation Schedule'
        merged_cellmodb131 = Model['B131']
        merged_cellmodb131.alignment = Alignment(horizontal='center', vertical='center')
        merged_cellmodb131.font = Font( size=18, bold=False)
        merged_cellmodb132 = Model['B132']
        merged_cellmodb132.alignment = Alignment(horizontal='center', vertical='center')
        merged_cellmodb132.font = Font( size=14, bold=False)
        
        Model['B133'] = '($)'
        Model['M134'] = 'Projected'
        Model['H135'] = '=K6-3'
        Model['I135'] = '=K6-2'
        Model['J135'] = '=K6-1'
        Model['K135'] = '=Scenarios!G6'
        Model['L135'] = '=Scenarios!H6'
        Model['M135'] = '=Scenarios!I6'
        Model['N135'] = '=Scenarios!J6'
        Model['O135'] = '=Scenarios!K6'
        
        Model['B135'] = 'Years Remaining Existing Assets:'
        Model['B136'] = 'Depreciation Years on New Assets:'
        Model['B138'] = 'Depreciation to Existing Assets'
        Model['F135'] = '=Assumptions!M24'
        Model['F136'] = '=Assumptions!M25'
        Model['K138'] = '=J71/F135'
        Model['L138'] = '=K138'
        Model['M138'] = '=L138'
        Model['N138'] = '=M138'
        Model['O138'] = '=N138'
        
        Model['D139'] = 'CAPEX'
        Modeld139 = Model['D139']
        Modeld139.font = Font(bold=True)
        
        Model['D140'] = '=K135'
        Model['D141'] = '=L135'
        Model['D142'] = '=M135'
        Model['D143'] = '=N135'
        Model['D144'] = '=O135'
        Model['B145'] = 'Total Depreciation'
        Modeld145 = Model['B145']
        Modeld145.font = Font(bold=True)
        
        Model['E140'] = '=-HLOOKUP(D140,$K$104:$O$114,ROWS($K$104:$K$114), FALSE)'
        Model['E141'] = '=-HLOOKUP(D141,$K$104:$O$114,ROWS($K$104:$K$114), FALSE)'
        Model['E142'] = '=-HLOOKUP(D142,$K$104:$O$114,ROWS($K$104:$K$114), FALSE)'
        Model['E143'] = '=-HLOOKUP(D143,$K$104:$O$114,ROWS($K$104:$K$114), FALSE)'
        Model['E144'] = '=-HLOOKUP(D144,$K$104:$O$114,ROWS($K$104:$K$114), FALSE)'
        
        Model['K140'] = '=IF($D$140 = $K$135, $E$140/$F$136/2,($E$140/$F$136))'
        Model['L140'] = '=IF($D$140 = $L$135, $E$140/$F$136/2,($E$140/$F$136))'
        Model['M140'] = '=IF($D$140 = $M$135, $E$140/$F$136/2,($E$140/$F$136))'
        Model['N140'] = '=IF($D$140 = $N$135, $E$140/$F$136/2,($E$140/$F$136))'
        Model['O140'] = '=IF($D$140 = $O$135, $E$140/$F$136/2,($E$140/$F$136))'
        
        Model['L141'] = '=IF($D$141 = $L$135, $E$141/$F$136/2,($E$141/$F$136))'
        Model['M141'] = '=IF($D$141 = $M$135, $E$141/$F$136/2,($E$141/$F$136))'
        Model['N141'] = '=IF($D$141 = $N$135, $E$141/$F$136/2,($E$141/$F$136))'
        Model['O141'] = '=IF($D$141 = $O$135, $E$141/$F$136/2,($E$141/$F$136))'
        
        Model['M142'] = '=IF($D$142 = $M$135, $E$142/$F$136/2,($E$142/$F$136))'
        Model['N142'] = '=IF($D$142 = $N$135, $E$142/$F$136/2,($E$142/$F$136))'
        Model['O142'] = '=IF($D$142 = $O$135, $E$142/$F$136/2,($E$142/$F$136))'
        
        Model['N143'] = '=IF($D$143 = $N$135, $E$143/$F$136/2,($E$143/$F$136))'
        Model['O143'] = '=IF($D$143 = $O$135, $E$143/$F$136/2,($E$143/$F$136))'
        
        Model['O144'] = '=IF($D$144 = $O$135, $E$144/$F$136/2,($E$144/$F$136))'
        
        Model['K145'] = '=SUM(K138:K144)'
        Model['L145'] = '=SUM(L138:L144)'
        Model['M145'] = '=SUM(M138:M144)'
        Model['N145'] = '=SUM(N138:N144)'
        Model['O145'] = '=SUM(O138:O144)'
        
        Model['O148'] = '="ECONOMIC: "&UPPER(CHOOSE(Scenarios!$D$6,Scenarios!C14,Scenarios!C15,Scenarios!C16))&" SCENARIO | SALES: "&UPPER(CHOOSE(Scenarios!$D$39,Scenarios!C14,Scenarios!C15,Scenarios!C16))&" SCENARIO | POLITICAL: "&UPPER(CHOOSE(Scenarios!$D$26,Scenarios!C14,Scenarios!C15,Scenarios!C16))&" SCENARIO"'
        Model['O148'].alignment = Alignment(horizontal='right')
        
        Model.merge_cells('B149:O149')
        Model['B149'] = company_name
        Model.merge_cells('B150:O150')
        Model['B150'] = 'Tax Schedule'
        merged_cellmodb149 = Model['B149']
        merged_cellmodb149.alignment = Alignment(horizontal='center', vertical='center')
        merged_cellmodb149.font = Font( size=18, bold=False)
        merged_cellmodb150 = Model['B150']
        merged_cellmodb150.alignment = Alignment(horizontal='center', vertical='center')
        merged_cellmodb150.font = Font( size=14, bold=False)
        Model['B151'] = '($)'
        Model['M152'] = 'Projected'
        Model['K153'] = '=Scenarios!G6'
        Model['L153'] = '=Scenarios!H6'
        Model['M153'] = '=Scenarios!I6'
        Model['N153'] = '=Scenarios!J6'
        Model['O153'] = '=Scenarios!K6'
        
        Model['C153'] = 'Tax Rate'
        modelc153 = Model['C153']
        modelc153.font = Font(bold=True)
        Model['E153'] = '=Scenarios!G31'
        Model['D156'] = 'Accounting EBT (as is on I/S)'
        Model['K156'] = '=K47'
        Model['L156'] = '=L47'
        Model['M156'] = '=M47'
        Model['N156'] = '=N47'
        Model['O156'] = '=O47'
        
        Model['D158'] = '=CONCATENATE("Income Taxes (",$E$153*100,"% of EBT)")'
        Model['K158'] = '=$E$153*K156'
        Model['L158'] = '=$E$153*L156'
        Model['M158'] = '=$E$153*M156'
        Model['N158'] = '=$E$153*N156'
        Model['O158'] = '=$E$153*O156'
        
        Model['O161'] = '="ECONOMIC: "&UPPER(CHOOSE(Scenarios!$D$6,Scenarios!C14,Scenarios!C15,Scenarios!C16))&" SCENARIO | SALES: "&UPPER(CHOOSE(Scenarios!$D$39,Scenarios!C14,Scenarios!C15,Scenarios!C16))&" SCENARIO | POLITICAL: "&UPPER(CHOOSE(Scenarios!$D$26,Scenarios!C14,Scenarios!C15,Scenarios!C16))&" SCENARIO"'
        Model['O161'].alignment = Alignment(horizontal='right')
        
        Model.merge_cells('B162:O162')
        Model['B162'] = company_name
        Model.merge_cells('B163:O163')
        Model['B163'] = 'Working Capital Schedule'
        merged_cellmodb162 = Model['B162']
        merged_cellmodb162.alignment = Alignment(horizontal='center', vertical='center')
        merged_cellmodb162.font = Font( size=18, bold=False)
        merged_cellmodb163 = Model['B163']
        merged_cellmodb163.alignment = Alignment(horizontal='center', vertical='center')
        merged_cellmodb163.font = Font( size=14, bold=False)
        Model['B164'] = '($)'
        Model['M165'] = 'Projected'
        Model['H166'] = '=K6-3'
        Model['I166'] = '=K6-2'
        Model['J166'] = '=K6-1'
        Model['K166'] = '=Scenarios!G6'
        Model['L166'] = '=Scenarios!H6'
        Model['M166'] = '=Scenarios!I6'
        Model['N166'] = '=Scenarios!J6'
        Model['O166'] = '=Scenarios!K6'
        
        Model['B168'] = 'Days per Year'

        Model['I168'] = '=DATE(I166,12,31)-DATE(H166,12,31)'
        Model['J168'] = '=DATE(J166,12,31)-DATE(I166,12,31)'
        Model['K168'] = '=DATE(K166,12,31)-DATE(J166,12,31)'
        Model['L168'] = '=DATE(L166,12,31)-DATE(K166,12,31)'
        Model['M168'] = '=DATE(M166,12,31)-DATE(L166,12,31)'
        Model['N168'] = '=DATE(N166,12,31)-DATE(M166,12,31)'
        Model['O168'] = '=DATE(O166,12,31)-DATE(N166,12,31)'
        Model['F168'] = '(Days)'
        
        Model['B170'] = 'Income Statement Items'
        modelb170 = Model['B170']
        modelb170.font = Font(bold=True)
        Model['C171'] = 'Revenue'
        Model['G171'] = '($ MM)'
        Model['C172'] = 'Cost of Revenue'
        Model['G172'] = '($ MM)'
        
        Model['I171'] = '=I30'
        Model['J171'] = '=J30'
        Model['K171'] = '=K30'
        Model['L171'] = '=L30'
        Model['M171'] = '=M30'
        Model['N171'] = '=N30'
        Model['O171'] = '=O30'
        
        Model['I172'] = '=I31'
        Model['J172'] = '=J31'
        Model['K172'] = '=K31'
        Model['L172'] = '=L31'
        Model['M172'] = '=M31'
        Model['N172'] = '=N31'
        Model['O172'] = '=O31'
        
        Model['I176'] = '=I68'
        Model['J176'] = '=J68'
        Model['K176'] = '=Assumptions!J35'
        Model['L176'] = '=Assumptions!K35'
        Model['M176'] = '=Assumptions!L35'
        Model['N176'] = '=Assumptions!M35'
        Model['O176'] = '=Assumptions!N35'
        
        Model['I175'] = '=I65'
        Model['J175'] = '=J65'
        Model['K175'] = '=Assumptions!J34'
        Model['L175'] = '=Assumptions!K34'
        Model['M175'] = '=Assumptions!L34'
        Model['N175'] = '=Assumptions!M34'
        Model['O175'] = '=Assumptions!N34'
        
        Model['I177'] = '=I66'
        Model['J177'] = '=J66'
        Model['K177'] = '=Assumptions!J36'
        Model['L177'] = '=Assumptions!K36'
        Model['M177'] = '=Assumptions!L36'
        Model['N177'] = '=Assumptions!M36'
        Model['O177'] = '=Assumptions!N36'
        
        Model['I178'] = '=I78'
        Model['J178'] = '=J78'
        Model['K178'] = '=Assumptions!J37'
        Model['L178'] = '=Assumptions!K37'
        Model['M178'] = '=Assumptions!L37'
        Model['N178'] = '=Assumptions!M37'
        Model['O178'] = '=Assumptions!N37'
        
        Model['I179'] = '=I84'
        Model['J179'] = '=J84'
        Model['K179'] = '=Assumptions!J38'
        Model['L179'] = '=Assumptions!K38'
        Model['M179'] = '=Assumptions!L38'
        Model['N179'] = '=Assumptions!M38'
        Model['O179'] = '=Assumptions!N38'
        
        Model['B174'] = 'Account Balances'
        modelb174 = Model['B174']
        modelb174.font = Font(bold=True)
        Model['C175'] = '=Assumptions!C34'
        Model['G175'] = '($ MM)'
        Model['C176'] = '=Assumptions!C35'
        Model['G176'] = '($ MM)'
        Model['C177'] = '=Assumptions!C36'
        Model['G177'] = '($ MM)'
        Model['C178'] = '=Assumptions!C37'
        Model['G178'] = '($ MM)'
        Model['C179'] = '=Assumptions!C38'
        Model['G179'] = '($ MM)'
        
        Model['C180'] = 'Net Working Capital'
        ModelC180 = Model['C180']
        ModelC180.font = Font(bold=True)
        Model['G180'] = '($ MM)'
        ModelG180 = Model['G180']
        ModelG180.font = Font(bold=True)
        
        Model['C182'] = 'Change in Net Working Capital'
        ModelC182 = Model['C182']
        ModelC182.font = Font(bold=True)
        Model['G182'] = '($ MM)'
        ModelG182 = Model['G182']
        ModelG182.font = Font(bold=True)
        
        Model['I180'] = '=SUM(I175:I177)-SUM(I178:I179)'
        Model['J180'] = '=SUM(J175:I177)-SUM(J178:J179)'
        Model['K180'] = '=SUM(K175:I177)-SUM(K178:K179)'
        Model['L180'] = '=SUM(L175:I177)-SUM(L178:L179)'
        Model['M180'] = '=SUM(M175:I177)-SUM(M178:M179)'
        Model['N180'] = '=SUM(N175:I177)-SUM(N178:N179)'
        Model['O180'] = '=SUM(O175:I177)-SUM(O178:O179)'
        
        Model['J182'] = '=I180-J180'
        Model['K182'] = '=J180-K180'
        Model['L182'] = '=K180-L180'
        Model['M182'] = '=L180-M180'
        Model['N182'] = '=M180-N180'
        Model['O182'] = '=N180-O180'
        
        Model['O185'] = '="ECONOMIC: "&UPPER(CHOOSE(Scenarios!$D$6,Scenarios!C14,Scenarios!C15,Scenarios!C16))&" SCENARIO | SALES: "&UPPER(CHOOSE(Scenarios!$D$39,Scenarios!C14,Scenarios!C15,Scenarios!C16))&" SCENARIO | POLITICAL: "&UPPER(CHOOSE(Scenarios!$D$26,Scenarios!C14,Scenarios!C15,Scenarios!C16))&" SCENARIO"'
        Model['O185'].alignment = Alignment(horizontal='right')
        
        Model.merge_cells('B186:O186')
        Model['B186'] = company_name
        Model.merge_cells('B187:O187')
        Model['B187'] = 'Debt & Interest Schedule'
        merged_cellmodB186 = Model['B186']
        merged_cellmodB186.alignment = Alignment(horizontal='center', vertical='center')
        merged_cellmodB186.font = Font( size=18, bold=False)
        merged_cellmodB187 = Model['B187']
        merged_cellmodB187.alignment = Alignment(horizontal='center', vertical='center')
        merged_cellmodB187.font = Font( size=14, bold=False)
        Model['B188'] = '($)'
        Model['M189'] = 'Projected'
        Model['J190'] = '=K6-1'
        Model['K190'] = '=Scenarios!G6'
        Model['L190'] = '=Scenarios!H6'
        Model['M190'] = '=Scenarios!I6'
        Model['N190'] = '=Scenarios!J6'
        Model['O190'] = '=Scenarios!K6'
        
        Model['B191'] = 'FINANCING COMPONENT'
        modelB191 = Model['B191']
        modelB191.font = Font(size=14, bold=True)
        Model['B193'] = 'Cash & Cash Equivalents'
        modelB193 = Model['B193']
        modelB193.font = Font(bold=True)
        
        Model['C194'] = 'Amount Outstanding Start'
        Model['C195'] = 'Change in Cash & Cash Equivalents'
        Model['C196'] = 'Amount Outstanding End'
        
        Model['C198'] = 'Interest Rate'
        Model['C199'] = 'Annual Interest Income'
        modelC199 = Model['C199']
        modelC199.font = Font(bold=True)
        
        Model['K198'] = '=Scenarios!G12'
        Model['L198'] = '=Scenarios!H12'
        Model['M198'] = '=Scenarios!I12'
        Model['N198'] = '=Scenarios!J12'
        Model['O198'] = '=Scenarios!K12'
        
        Model['K194'] = '=J196'
        Model['L194'] = '=K196'
        Model['M194'] = '=L196'
        Model['N194'] = '=M196'
        Model['O194'] = '=N196'
        
        Model['K195'] = '=K125'
        Model['L195'] = '=L125'
        Model['M195'] = '=M125'
        Model['N195'] = '=N125'
        Model['O195'] = '=O125'
        
        Model['J196'] = '=J67'
        Model['K196'] = '=SUM(K194:K195)'
        Model['L196'] = '=SUM(L194:L195)'
        Model['M196'] = '=SUM(M194:M195)'
        Model['N196'] = '=SUM(N194:N195)'
        Model['O196'] = '=SUM(O194:O195)'
        
        Model['K199'] = '=K198*K194'
        Model['L199'] = '=L198*L194'
        Model['M199'] = '=M198*M194'
        Model['N199'] = '=N198*N194'
        Model['O199'] = '=O198*O194'
        
        Model['B201'] = 'Long Term Debt'
        modelB201 = Model['B201']
        modelB201.font = Font(bold=True)
        Model['C202'] = 'Amount Outstanding Start'
        Model['C203'] = 'Additions / (Repayments)'
        Model['C204'] = 'Amount Outstanding End'
        
        Model['C206'] = 'Interest Rate'
        Model['C207'] = 'Annual Interest Expense'
        modelc207 = Model['C207']
        modelc207.font = Font(bold=True)
        
        Model['K203'] = '=Scenarios!G12'
        Model['L203'] = '=Scenarios!H12'
        Model['M203'] = '=Scenarios!I12'
        Model['N203'] = '=Scenarios!J12'
        Model['O203'] = '=Scenarios!K12'
        
        Model['J204'] = '=J83'
        
        Model['K202'] = '=J204'
        Model['L202'] = '=K204'
        Model['M202'] = '=L204'
        Model['N202'] = '=M204'
        Model['O202'] = '=N204'
        
        Model['K203'] = '=Assumptions!J41'
        Model['L203'] = '=Assumptions!K41'
        Model['M203'] = '=Assumptions!L41'
        Model['N203'] = '=Assumptions!M41'
        Model['O203'] = '=Assumptions!N41'
        
        Model['K204'] = '=SUM(K202:K203)'
        Model['L204'] = '=SUM(L202:L203)'
        Model['M204'] = '=SUM(M202:M203)'
        Model['N204'] = '=SUM(N202:N203)'
        Model['O204'] = '=SUM(O202:O203)'
        
        Model['K206'] = '=Scenarios!G12'
        Model['L206'] = '=Scenarios!H12'
        Model['M206'] = '=Scenarios!I12'
        Model['N206'] = '=Scenarios!J12'
        Model['O206'] = '=Scenarios!K12'
        
        Model['K207'] = '=K206*K202'
        Model['L207'] = '=L206*L202'
        Model['M207'] = '=M206*M202'
        Model['N207'] = '=N206*N202'
        Model['O207'] = '=O206*O202'
        
        Model['B209'] = 'Revolver'
        modelB209 = Model['B209']
        modelB209.font = Font(bold=True)
        Model['C210'] = 'Operating Cashflow'
        
        Model['K210'] = '=K110'
        Model['L210'] = '=L110'
        Model['M210'] = '=M110'
        Model['N210'] = '=N110'
        Model['O210'] = '=O110'
        
        Model['C211'] = 'Investing Cashflow'
        
        Model['K211'] = '=K115'
        Model['L211'] = '=L115'
        Model['M211'] = '=M115'
        Model['N211'] = '=N115'
        Model['O211'] = '=O115'
        
        Model['C212'] = 'Long Term Debt Repayments'
        
        Model['K212'] = '=K203'
        Model['L212'] = '=L203'
        Model['M212'] = '=M203'
        Model['N212'] = '=N203'
        Model['O212'] = '=O203'
        
        Model['C213'] = 'Common Stock Issuance / (Buy-back)'
        
        Model['K213'] = '=Assumptions!J42'
        Model['L213'] = '=Assumptions!K42'
        Model['M213'] = '=Assumptions!L42'
        Model['N213'] = '=Assumptions!M42'
        Model['O213'] = '=Assumptions!N42'
        
        Model['C214'] = 'Dividends Paid'
        
        Model['K214'] = '=-(K242+K251)'
        Model['L214'] = '=-(L242+L251)'
        Model['M214'] = '=-(M242+M251)'
        Model['N214'] = '=-(N242+N251)'
        Model['O214'] = '=-(O242+O251)'
        
        Model['C215'] = 'FCF After Mandatory Debt Repayment and Dividend'
        modelC215 = Model['C215']
        modelC215.font = Font(bold=True)
        
        Model['K215'] = '=SUM(K210:K214)'
        Model['L215'] = '=SUM(L210:L214)'
        Model['M215'] = '=SUM(M210:M214)'
        Model['N215'] = '=SUM(N210:N214)'
        Model['O215'] = '=SUM(O210:O214)'
        
        Model['C217'] = 'Revolver Outstanding Start'
        
        Model['K217'] = '=J219'
        Model['L217'] = '=K219'
        Model['M217'] = '=L219'
        Model['N217'] = '=M219'
        Model['O217'] = '=N219'
        
        Model['C218'] = 'Additions / (Repayments)'
        
        Model['K218'] = '=-MIN((K215+K194),K217)'
        Model['L218'] = '=-MIN((L215+L194),L217)'
        Model['M218'] = '=-MIN((M215+M194),M217)'
        Model['N218'] = '=-MIN((N215+N194),N217)'
        Model['O218'] = '=-MIN((O215+O194),O27)'
        
        Model['C219'] = 'Revolver Outstanding End'
        
        Model['K219'] = '=SUM(K217:K218)'
        Model['L219'] = '=SUM(L217:L218)'
        Model['M219'] = '=SUM(M217:M218)'
        Model['N219'] = '=SUM(N217:N218)'
        Model['O219'] = '=SUM(O217:O218)'
        
        Model['J219'] = '=J79'
        
        Model['C221'] = 'Interest Rate'
        
        Model['K221'] = '=Scenarios!G12'
        Model['L221'] = '=Scenarios!H12'
        Model['M221'] = '=Scenarios!I12'
        Model['N221'] = '=Scenarios!J12'
        Model['O221'] = '=Scenarios!K12'
        
        Model['C222'] = 'Annual Interest Expense'
        modelc222 = Model['C222']
        modelc222.font = Font(bold=True)
        
        Model['K222'] = '=K221*K217'
        Model['L222'] = '=L221*L217'
        Model['M222'] = '=M221*M217'
        Model['N222'] = '=N221*N217'
        Model['O222'] = '=O221*O217'
        
        Model['B224'] = 'Net Interest Expense'
        modelb224 = Model['B224']
        modelb224.font = Font(bold=True)
        
        Model['K224'] = '=K222+K207-K199'
        Model['L224'] = '=L222+L207-L199'
        Model['M224'] = '=M222+M207-M199'
        Model['N224'] = '=N222+N207-N199'
        Model['O224'] = '=O222+O207-O199'
        
        Model['O227'] = '="ECONOMIC: "&UPPER(CHOOSE(Scenarios!$D$6,Scenarios!C14,Scenarios!C15,Scenarios!C16))&" SCENARIO | SALES: "&UPPER(CHOOSE(Scenarios!$D$39,Scenarios!C14,Scenarios!C15,Scenarios!C16))&" SCENARIO | POLITICAL: "&UPPER(CHOOSE(Scenarios!$D$26,Scenarios!C14,Scenarios!C15,Scenarios!C16))&" SCENARIO"'
        Model['O227'].alignment = Alignment(horizontal='right')
        
        Model.merge_cells('B228:O228')
        Model['B228'] = company_name
        Model.merge_cells('B229:O229')
        Model['B229'] = 'Shareholder Equity Schedule'
        merged_cellmodb228 = Model['B228']
        merged_cellmodb228.alignment = Alignment(horizontal='center', vertical='center')
        merged_cellmodb228.font = Font( size=18, bold=False)
        merged_cellmodb229 = Model['B229']
        merged_cellmodb229.alignment = Alignment(horizontal='center', vertical='center')
        merged_cellmodb229.font = Font( size=14, bold=False)
        Model['B230'] = '($)'
        Model['M231'] = 'Projected'
        Model['J232'] = '=K6-1'
        Model['K232'] = '=Scenarios!G6'
        Model['L232'] = '=Scenarios!H6'
        Model['M232'] = '=Scenarios!I6'
        Model['N232'] = '=Scenarios!J6'
        Model['O232'] = '=Scenarios!K6'
        
        Model['B235'] = 'Common Shares'
        modelb235 = Model['B235']
        modelb235.font = Font(bold=True)
        
        Model['C236'] = 'Amount Outstanding Start'
        
        Model['K236'] = '=J238'
        Model['L236'] = '=K238'
        Model['M236'] = '=L238'
        Model['N236'] = '=M238'
        Model['O236'] = '=N238'
        
        Model['C237'] = 'New Issuance / (Buy-back)'
        
        Model['K237'] = '=Assumptions!J42'
        Model['L237'] = '=Assumptions!K42'
        Model['M237'] = '=Assumptions!L42'
        Model['N237'] = '=Assumptions!M42'
        Model['O237'] = '=Assumptions!N42'
        
        Model['C238'] = 'Amount Outstanding End'
        
        Model['J238'] = '=J90'
        
        Model['K238'] = '=SUM(K236:K237)'
        Model['L238'] = '=SUM(L236:L237)'
        Model['M238'] = '=SUM(M236:M237)'
        Model['N238'] = '=SUM(N236:N237)'
        Model['O238'] = '=SUM(O236:O237)'
        
        Model['C240'] = 'Dividend Payout Rate'
        
        Model['K240'] = '=Assumptions!M6'
        Model['L240'] = '=Assumptions!M6'
        Model['M240'] = '=Assumptions!M6'
        Model['N240'] = '=Assumptions!M6'
        Model['O240'] = '=Assumptions!M6'
        
        Model['C241'] = 'Net Income'
        
        Model['K241'] = '=J50'
        Model['L241'] = '=K50'
        Model['M241'] = '=L50'
        Model['N241'] = '=M50'
        Model['O241'] = '=N50'
        
        Model['C242'] = 'Common Dividend'
        modelC242 = Model['C242']
        modelC242.font = Font(bold=True)
        Model['K242'] = '=MAX(K241*K240,0)'
        Model['L242'] = '=MAX(L241*L240,0)'
        Model['M242'] = '=MAX(M241*M240,0)'
        Model['N242'] = '=MAX(N241*N240,0)'
        Model['O242'] = '=MAX(O241*O240,0)'
        
        Model['B244'] = 'Preferred Shares'
        modelb244 = Model['B244']
        modelb244.font = Font(bold=True)
        
        Model['C245'] = 'Amount Outstanding Start'
        
        Model['K245'] = '=J247'
        Model['L245'] = '=K247'
        Model['M245'] = '=L247'
        Model['N245'] = '=M247'
        Model['O245'] = '=N247'
        
        Model['C246'] = 'Change in Preferred Shares'
        
        Model['K246'] = '=Assumptions!J43'
        Model['L246'] = '=Assumptions!K43'
        Model['M246'] = '=Assumptions!L43'
        Model['N246'] = '=Assumptions!M43'
        Model['O246'] = '=Assumptions!N43'
        
        Model['C247'] = 'Amount Outstanding End'
        
        Model['J247'] = '=J98'
        Model['K247'] = '=SUM(K245:K246)'
        Model['L247'] = '=SUM(L245:L246)'
        Model['M247'] = '=SUM(M245:M246)'
        Model['N247'] = '=SUM(N245:N246)'
        Model['O247'] = '=SUM(O245:O246)'
        
        Model['C249'] = 'Dividend Payout Rate'
        
        Model['K249'] = '=Assumptions!M7'
        Model['L249'] = '=Assumptions!M7'
        Model['M249'] = '=Assumptions!M7'
        Model['N249'] = '=Assumptions!M7'
        Model['O249'] = '=Assumptions!M7'
        
        Model['C250'] = 'Net Income'
        
        Model['K250'] = '=J50'
        Model['L250'] = '=K50'
        Model['M250'] = '=L50'
        Model['N250'] = '=M50'
        Model['O250'] = '=N50'
        
        Model['C251'] = 'Preferred Dividend'
        modelC251 = Model['C251']
        modelC251.font = Font(bold=True)
        
        Model['K251'] = '=MAX(K249*K250,0)'
        Model['L251'] = '=MAX(L249*L250,0)'
        Model['M251'] = '=MAX(M249*M250,0)'
        Model['N251'] = '=MAX(N249*N250,0)'
        Model['O251'] = '=MAX(O249*O250,0)'
        
        Model['B253'] = 'Accumulated Other Comprehensive Income (Loss)'
        modelb253 = Model['B253']
        modelb253.font = Font(bold=True)
        
        Model['C254'] = 'Amount Outstanding Start'
        
        Model['K254'] = '=J256'
        Model['L254'] = '=K256'
        Model['M254'] = '=L256'
        Model['N254'] = '=M256'
        Model['O254'] = '=N256'
        
        Model['C255'] = 'Change in Other Comprehensive Income (Loss)'
        
        Model['K255'] = '=Assumptions!J44'
        Model['L255'] = '=Assumptions!K44'
        Model['M255'] = '=Assumptions!L44'
        Model['N255'] = '=Assumptions!M44'
        Model['O255'] = '=Assumptions!N44'
        
        Model['C256'] = 'Amount Outstanding End'
        modelc256 = Model['C256']
        modelc256.font = Font(bold=True)
        
        Model['J256'] = '=J91'
        Model['K256'] = '=SUM(K254:K255)'
        Model['L256'] = '=SUM(L254:L255)'
        Model['M256'] = '=SUM(M254:M255)'
        Model['N256'] = '=SUM(N254:N255)'
        Model['O256'] = '=SUM(O254:O255)'
        
        Model['B258'] = 'Other Stockholder Equity'
        modelb258 = Model['B258']
        modelb258.font = Font(bold=True)
        
        Model['C259'] = 'Amount Outstanding Start'
        
        Model['K259'] = '=J261'
        Model['L259'] = '=K261'
        Model['M259'] = '=L261'
        Model['N259'] = '=M261'
        Model['O259'] = '=N261'
        
        Model['C260'] = 'Change in Other Stockholder Equity'
    
        Model['K260'] = '=Assumptions!J48'
        Model['L260'] = '=Assumptions!K48'
        Model['M260'] = '=Assumptions!L48'
        Model['N260'] = '=Assumptions!M48'
        Model['O260'] = '=Assumptions!N48'
        
        Model['C261'] = 'Amount Outstanding End'
        modelC261 = Model['C261']
        modelC261.font = Font(bold=True)
        
        Model['J261'] = '=J93'
        Model['K261'] = '=SUM(K259:K260)'
        Model['L261'] = '=SUM(L259:L260)'
        Model['M261'] = '=SUM(M259:M260)'
        Model['N261'] = '=SUM(N259:N260)'
        Model['O261'] = '=SUM(O259:O260)'
        
        Model['B263'] = 'Retained Earnings'
        modelb263 = Model['B263']
        modelb263.font = Font(bold=True)
        
        Model['C264'] = 'Amount Outstanding Start'
        
        Model['K264'] = '=J268'
        Model['L264'] = '=K268'
        Model['M264'] = '=L268'
        Model['N264'] = '=M268'
        Model['O264'] = '=N268'
        
        Model['C265'] = 'Net Income'
        
        Model['K265'] = '=K50'
        Model['L265'] = '=L50'
        Model['M265'] = '=M50'
        Model['N265'] = '=N50'
        Model['O265'] = '=O50'
        
        Model['C266'] = 'Common Dividend'
        
        Model['K266'] = '=-K242'
        Model['L266'] = '=-L242'
        Model['M266'] = '=-M242'
        Model['N266'] = '=-N242'
        Model['O266'] = '=-O242'
        
        Model['C267'] = 'Preferred Dividend'
        
        Model['K267'] = '=-K251'
        Model['L267'] = '=-L251'
        Model['M267'] = '=-M251'
        Model['N267'] = '=-N251'
        Model['O267'] = '=-O251'
        
        Model['C268'] = 'Amount Outstanding End'
        modelb268 = Model['B268']
        modelb268.font = Font(bold=True)
        
        Model['J268'] = '=J93'
        Model['K268'] = '=SUM(K264:K267)'
        Model['L268'] = '=SUM(L264:L267)'
        Model['M268'] = '=SUM(M264:M267)'
        Model['N268'] = '=SUM(N264:N267)'
        Model['O268'] = '=SUM(O264:O267)'
        
        Model['O271'] = '="ECONOMIC: "&UPPER(CHOOSE(Scenarios!$D$6,Scenarios!C14,Scenarios!C15,Scenarios!C16))&" SCENARIO | SALES: "&UPPER(CHOOSE(Scenarios!$D$39,Scenarios!C14,Scenarios!C15,Scenarios!C16))&" SCENARIO | POLITICAL: "&UPPER(CHOOSE(Scenarios!$D$26,Scenarios!C14,Scenarios!C15,Scenarios!C16))&" SCENARIO"'
        Model['O271'].alignment = Alignment(horizontal='right')
        
        Model.merge_cells('B272:O272')
        Model['B272'] = company_name
        Model.merge_cells('B273:O273')
        Model['B273'] = 'DCF Valuation'
        merged_cellmodb272 = Model['B272']
        merged_cellmodb272.alignment = Alignment(horizontal='center', vertical='center')
        merged_cellmodb272.font = Font( size=18, bold=False)
        merged_cellmodb273 = Model['B273']
        merged_cellmodb273.alignment = Alignment(horizontal='center', vertical='center')
        merged_cellmodb273.font = Font( size=14, bold=False)
        Model['B274'] = '($)'
        Model['M275'] = 'Projected'
        #Model['J276'] = '=K6-1'
        Model['K276'] = '=Scenarios!G6'
        Model['L276'] = '=Scenarios!H6'
        Model['M276'] = '=Scenarios!I6'
        Model['N276'] = '=Scenarios!J6'
        Model['O276'] = '=Scenarios!K6'
        
        Model['C275'] = 'GDP Growth'
        modelc275 = Model['C275']
        modelc275.font = Font(bold=True)
        Model['E275'] = '=Scenarios!G18'
        
        Model['C276'] = 'Tax Rate'
        modelc276 = Model['C276']
        modelc276.font = Font(bold=True)
        Model['E276'] = '=Scenarios!G31'
        
        Model['B278'] = 'Unlevered Free Cash Flow'
        modelB278 = Model['B278']
        modelB278.font = Font(bold=True)
        Model['C279'] = 'EBIT'
        Model['K279'] = '=K44'
        Model['L279'] = '=L44'
        Model['M279'] = '=M44'
        Model['N279'] = '=N44'
        Model['O279'] = '=O44'
        
        Model['C280'] = 'Less Tax'
        Model['K280'] = '=(1-E276)*K279'
        Model['L280'] = '=(1-E276)*L279'
        Model['M280'] = '=(1-E276)*M279'
        Model['N280'] = '=(1-E276)*N279'
        Model['O280'] = '=(1-E276)*O279'
        
        Model['C281'] = '+ Depreciation & Amortization'
        Model['K281'] = '=K280+K43'
        Model['L281'] = '=L280+L43'
        Model['M281'] = '=M280+M43'
        Model['N281'] = '=N280+N43'
        Model['O281'] = '=O280+O43'
        
        Model['C282'] = 'Less CapEx'
        Model['K282'] = '=K281+K113'
        Model['L282'] = '=L281+L113'
        Model['M282'] = '=M281+M113'
        Model['N282'] = '=N281+N113'
        Model['O282'] = '=O281+O113'
        
        Model['C283'] = '(UFCF) Less Change in Net Working Capital'
        modelc283 = Model['C283']
        modelc283.font = Font(bold=True)
        Model['K283'] = '=K282-K182'
        Model['L283'] = '=L282-L182'
        Model['M283'] = '=M282-M182'
        Model['N283'] = '=N282-N182'
        Model['O283'] = '=O282-O182'
        
        Model['C285'] = 'Share Price'
        modelc285 = Model['C285']
        modelc285.font = Font(bold=True)
        Model['E285'] = '=Assumptions!M5'
        
        Model['C286'] = 'Shares Outstanding'
        modelc286 = Model['C286']
        modelc286.font = Font(bold=True)
        Model['E286'] = data.loc[data['calendarYear'] == one_pre, 'weightedAverageShsOut'].values[0]
        
        Model['C287'] = 'Market Value'
        modelc287 = Model['C287']
        modelc287.font = Font(bold=True)
        Model['E287'] = '=E286*E285'
        
        Model['C288'] = 'Total Debt'
        modelc288 = Model['C288']
        modelc288.font = Font(bold=True)
        Model['E288'] = data.loc[data['calendarYear'] == one_pre, 'totalDebt'].values[0]
        
        Model['C289'] = 'Enterprise Value'
        modelc289 = Model['C289']
        modelc289.font = Font(bold=True)
        Model['E289'] = '=E288 + E287 - J67'
        
        Model['C290'] = 'Cost of Debt'
        modelc290 = Model['C290']
        modelc290.font = Font(bold=True)
        Model['E290'] = '=J46/E288'
        
        Model['C291'] = 'β'
        modelc291 = Model['C291']
        modelc291.font = Font(bold=True)
        Model['E291'] = ''
        
        Model['C292'] = 'Rf'
        modelc292 = Model['C292']
        modelc292.font = Font(bold=True)
        Model['E292'] = ''
        
        Model['C293'] = 'Rm'
        modelc293 = Model['C293']
        modelc293.font = Font(bold=True)
        Model['E293'] = ''
        
        Model['C294'] = 'Cost of Equity'
        modelc294 = Model['C294']
        modelc294.font = Font(bold=True)
        Model['E294'] = '=E292+E291*(E293-E292)'
        
        Model['C295'] = 'WACC'
        modelc295 = Model['C295']
        modelc295.font = Font(bold=True)
        Model['E295'] = '=((E287/E289)*E294)+((E288/E289)*E290*(1-E276))'
        
        Model['C296'] = 'Terminal Value'
        modelc296 = Model['C296']
        modelc296.font = Font(bold=True)
        Model['E296'] = '=(O283*(1+E275))/(E295-E275)'
        
        Model['C298'] = 'UFCF NPV'
        modelc298 = Model['C298']
        modelc298.font = Font(bold=True)
        Model['E298'] = '=NPV(E295, K283:O283)'
        
        Model['C297'] = 'Terminal NPV'
        modelc297 = Model['C297']
        modelc297.font = Font(bold=True)
        Model['E297'] = '=NPV(E295, E296)'
        
        Model['C299'] = 'Total NPV'
        modelc299 = Model['C299']
        modelc299.font = Font(bold=True)
        Model['E299'] = '=E297+E298'
        
        Model['C300'] = 'Equity Value'
        modelc300 = Model['C300']
        modelc300.font = Font(bold=True)
        Model['E300'] = '=E299-E288'
        
        Model['C301'] = 'Forecast Price'
        modelc301 = Model['C300']
        modelc301.font = Font(bold=True)
        Model['E301'] = '=E300/E286'
        
        # Apply the thick bottom border to each cell in the range B25:O25
        for col in range(2, 16):  # Columns B (2) to O (15)
            cell = Model.cell(row=25, column=col)
            cell.border = thick_bottom_border  # Apply the thick border
        
        # Apply the thick bottom border to each cell in the range B60:O60
        for col in range(2, 16):  # Columns B (2) to O (15)
            cell = Model.cell(row=60, column=col)
            cell.border = thick_bottom_border  # Apply the thick border
            
        # Apply the thick bottom border to each cell in the range B102:O102
        for col in range(2, 16):  # Columns B (2) to O (15)
            cell = Model.cell(row=101, column=col)
            cell.border = thick_bottom_border  # Apply the thick border
            
        # Apply the thick bottom border to each cell in the range B132:O132
        for col in range(2, 16):  # Columns B (2) to O (15)
            cell = Model.cell(row=132, column=col)
            cell.border = thick_bottom_border  # Apply the thick border
        
        # Apply the thick bottom border to each cell in the range B150:O150
        for col in range(2, 16):  # Columns B (2) to O (15)
            cell = Model.cell(row=150, column=col)
            cell.border = thick_bottom_border  # Apply the thick border
            
        # Apply the thick bottom border to each cell in the range B163:O163
        for col in range(2, 16):  # Columns B (2) to O (15)
            cell = Model.cell(row=163, column=col)
            cell.border = thick_bottom_border  # Apply the thick border
            
        # Apply the thick bottom border to each cell in the range B187:O187
        for col in range(2, 16):  # Columns B (2) to O (15)
            cell = Model.cell(row=187, column=col)
            cell.border = thick_bottom_border  # Apply the thick border
            
        # Apply the thick bottom border to each cell in the range B229:O229
        for col in range(2, 16):  # Columns B (2) to O (15)
            cell = Model.cell(row=229, column=col)
            cell.border = thick_bottom_border  # Apply the thick border
            
        # Apply the thick bottom border to each cell in the range B273:O273
        for col in range(2, 16):  # Columns B (2) to O (15)
            cell = Model.cell(row=273, column=col)
            cell.border = thick_bottom_border  # Apply the thick border
        
        # Define a thin bottom border
        thin_bottom_border = Border(bottom=Side(style='thin'))

        # Apply the thick bottom border to each cell in the range J9:O9
        for col in range(10, 16):  # Columns J (10) to O (15)
            cell = Model.cell(row=9, column=col)
            cell.border = thin_bottom_border  # Apply the thin border
            
        # Apply the thick bottom border to each cell in the range J14:O14
        for col in range(10, 16):  # Columns J (10) to O (15)
            cell = Model.cell(row=14, column=col)
            cell.border = thin_bottom_border  # Apply the thin border
            
        # Apply the thick bottom border to each cell in the range J20:O20
        for col in range(10, 16):  # Columns J (10) to O (15)
            cell = Model.cell(row=19, column=col)
            cell.border = thin_bottom_border  # Apply the thin border
        
        # Define a thick bottom border
        thick_bottom_border = Border(bottom=Side(style='thick'))

        # Apply the thick bottom border to each cell in the range B3:O3
        for col in range(2, 16):  # Columns B (2) to O (15)
            cell = Model.cell(row=3, column=col)
            cell.border = thick_bottom_border  # Apply the thick border
        
        # Access the ASSUMPTIONS worksheet and set titles + subtitle
        Assumptions = writer.sheets['Assumptions']
        Assumptions.merge_cells('B1:O1')
        Assumptions['B1'] = company_name
        Assumptions.merge_cells('B2:O2')
        Assumptions['B2'] = 'Assumptions and Inputs'
        # Center the title and adjust font
        merged_cellassb1 = Assumptions['B1']
        merged_cellassb1.alignment = Alignment(horizontal='center', vertical='center')
        merged_cellassb1.font = Font( size=18, bold=False)
        # Center the section title and adjust font
        merged_cellassb2 = Assumptions['B2']
        merged_cellassb2.alignment = Alignment(horizontal='center', vertical='center')
        merged_cellassb2.font = Font( size=14, bold=False)
        # Define Assumptions main subtitles
        Assumptions['C4'] = 'General Overview'
        
        Assumptions['C9'] = 'Interest Rates'
        Assumptions['C10'] = 'Scenario Cases'
        Ass_Case1 = Assumptions['C10']
        Ass_Case1.font = Font(bold=True)
        Assumptions['C11'] = 'Best Case'
        Assumptions['C12'] = 'Base Case'
        Assumptions['C13'] = 'Worst Case'
        Assumptions.merge_cells('G10:H10')
        Assumptions['G10'] = '=Scenarios!G6&" - "&Scenarios!K6'
        Ass_IR_sub = Assumptions['G10'] 
        Ass_IR_sub.alignment = Alignment(horizontal='center', vertical='center')
        Assumptions.merge_cells('G11:H11')
        Assumptions['G11'] = '-0.25%'
        Ass_IR_best = Assumptions['G11'] 
        Ass_IR_best.alignment = Alignment(horizontal='center', vertical='center')
        Assumptions.merge_cells('G12:H12')
        Assumptions['G12'] = 'Research Forecast'
        Ass_IR_base = Assumptions['G12'] 
        Ass_IR_base.alignment = Alignment(horizontal='center', vertical='center')
        Assumptions.merge_cells('G13:H13')
        Assumptions['G13'] = '0.25%'
        Ass_IR_worst = Assumptions['G13'] 
        Ass_IR_worst.alignment = Alignment(horizontal='center', vertical='center')
        
        Assumptions['C15'] = 'GDP Growth'
        Assumptions['C16'] = 'Scenario Cases'
        Ass_Case1 = Assumptions['C16']
        Ass_Case1.font = Font(bold=True)
        Assumptions['C17'] = 'Best Case'
        Assumptions['C18'] = 'Base Case'
        Assumptions['C19'] = 'Worst Case'
        Assumptions.merge_cells('G16:H16')
        Assumptions['G16'] = '=Scenarios!G6&" - "&Scenarios!K6'
        Ass_IR_sub = Assumptions['G16'] 
        Ass_IR_sub.alignment = Alignment(horizontal='center', vertical='center')
        Assumptions.merge_cells('G17:H17')
        Assumptions['G17'] = '1%'
        Ass_IR_best = Assumptions['G17'] 
        Ass_IR_best.alignment = Alignment(horizontal='center', vertical='center')
        Assumptions.merge_cells('G18:H18')
        Assumptions['G18'] = 'Research Forecast'
        Ass_IR_base = Assumptions['G18'] 
        Ass_IR_base.alignment = Alignment(horizontal='center', vertical='center')
        Assumptions.merge_cells('G19:H19')
        Assumptions['G19'] = '-1%'
        Ass_IR_worst = Assumptions['G19'] 
        Ass_IR_worst.alignment = Alignment(horizontal='center', vertical='center')
        
        Assumptions['J4'] = 'Equity'
        Assumptions['J5'] = 'Stock Price as Of:'
        Assumptions['L5'] = 'DATE'
        Assumptions['J6'] = 'Common Dividend Pay Rate'
        Assumptions['J7'] = 'Preferred Dividend Pay Rate'
        Assumptions.merge_cells('M5:N5')
        Assumptions.merge_cells('M6:N6')
        Assumptions.merge_cells('M7:N7')
        Assumptions['M5'] = '122'
        assm5 = Assumptions['M5']
        assm5.alignment = Alignment(horizontal='center', vertical='center')
        Assumptions['M6'] = '1.02%'
        assm6 = Assumptions['M6']
        assm6.alignment = Alignment(horizontal='center', vertical='center')
        Assumptions['M7'] = '0%'
        assm7 = Assumptions['M7']
        assm7.alignment = Alignment(horizontal='center', vertical='center')
        
        Assumptions['J9'] = 'Tax Rate'
        Assumptions['J10'] = 'Scenario Cases'
        Ass_Case1 = Assumptions['J10']
        Ass_Case1.font = Font(bold=True)
        Assumptions['J11'] = 'Best Case'
        Assumptions['J12'] = 'Base Case'
        Assumptions['J13'] = 'Worst Case'
        Assumptions.merge_cells('M10:N10')
        Assumptions['M10'] = '=Scenarios!G6&" - "&Scenarios!K6'
        Ass_IR_sub = Assumptions['M10'] 
        Ass_IR_sub.alignment = Alignment(horizontal='center', vertical='center')
        Assumptions.merge_cells('M11:N11')
        Assumptions['M11'] = '16%'
        Ass_IR_best = Assumptions['M11'] 
        Ass_IR_best.alignment = Alignment(horizontal='center', vertical='center')
        Assumptions.merge_cells('M12:N12')
        Assumptions['M12'] = '21%'
        Ass_IR_base = Assumptions['M12'] 
        Ass_IR_base.alignment = Alignment(horizontal='center', vertical='center')
        Assumptions.merge_cells('M13:N13')
        Assumptions['M13'] = '25%'
        Ass_IR_worst = Assumptions['M13'] 
        Ass_IR_worst.alignment = Alignment(horizontal='center', vertical='center')
        
        Assumptions['J15'] = 'Segment A Growth'
        Assumptions['J16'] = 'Scenario Cases'
        Ass_Case1 = Assumptions['J16']
        Ass_Case1.font = Font(bold=True)
        Assumptions['J17'] = 'Best Case'
        Assumptions['J18'] = 'Base Case'
        Assumptions['J19'] = 'Worst Case'
        Assumptions.merge_cells('M16:N16')
        Assumptions['M16'] = '=Scenarios!G6&" - "&Scenarios!K6'
        Ass_IR_sub = Assumptions['M16'] 
        Ass_IR_sub.alignment = Alignment(horizontal='center', vertical='center')
        Assumptions.merge_cells('M17:N17')
        Assumptions['M17'] = '5%'
        Ass_IR_best = Assumptions['M17'] 
        Ass_IR_best.alignment = Alignment(horizontal='center', vertical='center')
        Assumptions.merge_cells('M18:N18')
        Assumptions['M18'] = 'Research Forecast'
        Ass_IR_base = Assumptions['M18'] 
        Ass_IR_base.alignment = Alignment(horizontal='center', vertical='center')
        Assumptions.merge_cells('M19:N19')
        Assumptions['M19'] = '-5%'
        Ass_IR_worst = Assumptions['M19'] 
        Ass_IR_worst.alignment = Alignment(horizontal='center', vertical='center')
        
        Assumptions['C21'] = 'Segment B Growth'
        Assumptions['C22'] = 'Scenario Cases'
        Ass_Case1 = Assumptions['C22']
        Ass_Case1.font = Font(bold=True)
        Assumptions['C23'] = 'Best Case'
        Assumptions['C24'] = 'Base Case'
        Assumptions['C25'] = 'Worst Case'
        Assumptions.merge_cells('G22:H22')
        Assumptions['G22'] = '=Scenarios!G6&" - "&Scenarios!K6'
        Ass_IR_sub = Assumptions['G22'] 
        Ass_IR_sub.alignment = Alignment(horizontal='center', vertical='center')
        Assumptions.merge_cells('G23:H23')
        Assumptions['G23'] = '7.5%'
        Ass_IR_best = Assumptions['G23'] 
        Ass_IR_best.alignment = Alignment(horizontal='center', vertical='center')
        Assumptions.merge_cells('G24:H24')
        Assumptions['G24'] = 'Research Forecast'
        Ass_IR_base = Assumptions['G24'] 
        Ass_IR_base.alignment = Alignment(horizontal='center', vertical='center')
        Assumptions.merge_cells('G25:H25')
        Assumptions['G25'] = '-7.5%'
        Ass_IR_worst = Assumptions['G25'] 
        Ass_IR_worst.alignment = Alignment(horizontal='center', vertical='center')
        
        Assumptions['J21'] = 'Depreciation'
        Assumptions['J23'] = 'Depreciation Method'
        Assumptions['J24'] = 'Existing Asset Lifespan'
        Assumptions['J25'] = 'New Asset Lifespan'
        
        Assumptions.merge_cells('M23:N23')
        assm23 = Assumptions['M23']
        assm23.alignment = Alignment(horizontal='center', vertical='center')
        Assumptions['M23'] = 'Straight Line'
        
        Assumptions.merge_cells('M24:N24')
        assm24 = Assumptions['M24']
        assm24.alignment = Alignment(horizontal='center', vertical='center')
        Assumptions['M24'] = '8'
        
        Assumptions.merge_cells('M25:N25')
        assm25 = Assumptions['M25']
        assm25.alignment = Alignment(horizontal='center', vertical='center')
        Assumptions['M25'] = '13'
        
        
        Assumptions['C27'] = 'Other Assumptions'
        
        Assumptions['J27'] = '=Scenarios!G6'
        Assumptions['K27'] = '=Scenarios!H6'
        Assumptions['L27'] = '=Scenarios!I6'
        Assumptions['M27'] = '=Scenarios!J6'
        Assumptions['N27'] = '=Scenarios!K6'
        
        Assumptions['C29'] = 'Investing Activities'
        assc29 = Assumptions['C29']
        assc29.font = Font(bold=True)
        
        Assumptions['C30'] = 'CAPEX'
        Assumptions['C31'] = 'Other Investing Activities'
        
        
        Assumptions['C33'] = 'Working Capital'
        assc33 = Assumptions['C33']
        assc33.font = Font(bold=True)
        Assumptions['C34'] = 'Accounts Receivable'
        Assumptions['C35'] = 'Other Current Assets'
        Assumptions['C36'] = 'Inventory'
        Assumptions['C37'] = 'Accounts Payable'
        Assumptions['C38'] = 'Other Current Liabilities'
        
        Assumptions['C40'] = 'Changes in Equity and Debt'
        assc40 = Assumptions['C40']
        assc40.font = Font(bold=True)
        Assumptions['C41'] = 'Long Term Debt Issuance / (Repayments)'
        
        Assumptions['C42'] = 'Common Stock Issuance / (Repayments)'
        
        Assumptions['J42'] = ''
        Assumptions['K42'] = ''
        Assumptions['L42'] = ''
        Assumptions['M42'] = ''
        Assumptions['N42'] = ''
        
        Assumptions['C43'] = 'Change in Preferred Shares'
        
        Assumptions['J43'] = ''
        Assumptions['K43'] = ''
        Assumptions['L43'] = ''
        Assumptions['M43'] = ''
        Assumptions['N43'] = ''
        
        Assumptions['C44'] = 'Change in Comprehensive Income (Loss)'
        
        Assumptions['J44'] = ''
        Assumptions['K44'] = ''
        Assumptions['L44'] = ''
        Assumptions['M44'] = ''
        Assumptions['N44'] = ''
        
        Assumptions['C46'] = 'Assets & Liabilities'
        assc46 = Assumptions['C46']
        assc46.font = Font(bold=True)
        
        Assumptions['C47'] = 'Other Stockholder Equity'
        
        Assumptions['J47'] = ''
        Assumptions['K47'] = ''
        Assumptions['L47'] = ''
        Assumptions['M47'] = ''
        Assumptions['N47'] = ''
        
        Assumptions['C48'] = 'Other Non-Current Liabilities'
        
        Assumptions['J48'] = ''
        Assumptions['K48'] = ''
        Assumptions['L48'] = ''
        Assumptions['M48'] = ''
        Assumptions['N48'] = ''
        
        Assumptions['C50'] = 'Operating Activity'
        assc50 = Assumptions['C50']
        assc50.font = Font(bold=True)
        Assumptions['C51'] = 'Other Non-Cash Items'
        Assumptions['J51'] = ''
        Assumptions['K51'] = ''
        Assumptions['L51'] = ''
        Assumptions['M51'] = ''
        Assumptions['N51'] = ''
        
        Assumptions['C53'] = 'Financing Activity'
        assc53 = Assumptions['C53']
        assc53.font = Font(bold=True)
        Assumptions['C54'] = 'Other Financing Activity'
        Assumptions['J54'] = ''
        Assumptions['K54'] = ''
        Assumptions['L54'] = ''
        Assumptions['M54'] = ''
        Assumptions['N54'] = ''
        
        Assumptions['C56'] = 'Cost of Production'
        assc56 = Assumptions['C56']
        assc56.font = Font(bold=True)
        Assumptions['C57'] = 'Cost of Revenue'
        Assumptions['J57'] = ''
        Assumptions['K57'] = ''
        Assumptions['L57'] = ''
        Assumptions['M57'] = ''
        Assumptions['N57'] = ''
        Assumptions['C58'] = 'Research & Development'
        Assumptions['J58'] = ''
        Assumptions['K58'] = ''
        Assumptions['L58'] = ''
        Assumptions['M58'] = ''
        Assumptions['N58'] = ''
        Assumptions['C59'] = 'S G & A'
        Assumptions['J59'] = ''
        Assumptions['K59'] = ''
        Assumptions['L59'] = ''
        Assumptions['M59'] = ''
        Assumptions['N59'] = ''
        
        # Define a thick bottom border
        thick_bottom_border = Border(bottom=Side(style='thick'))

        # Apply the thick bottom border to each cell in the range B2:O2
        for col in range(2, 16):  # Columns B (2) to O (15)
            cell = Assumptions.cell(row=2, column=col)
            cell.border = thick_bottom_border  # Apply the thick border
            
        # ACCESS SCENARIOS
        # Merge cells B1:O1 and set the title
        Scenarios.merge_cells('B1:O1')
        Scenarios['B1'] = company_name  
        # Merge cells B2:O2 and set subtitle
        Scenarios.merge_cells('B2:O2')
        Scenarios['B2'] = 'Economic and Sales Scenarios'
        
        
        # Set the years in the cells
        set_years_in_range(Scenarios, 'G6', 'K6', years)
        
        # Set switch
        Scenarios['B6'] = 'ECONOMIC SWITCH:'
        Scenarios['D6'] = '1'
        
        # Center the title and adjust font
        merged_cellb1 = Scenarios['B1']
        merged_cellb1.alignment = Alignment(horizontal='center', vertical='center')
        merged_cellb1.font = Font( size=18, bold=False)
        # Center the section title and adjust font
        merged_cellb2 = Scenarios['B2']
        merged_cellb2.alignment = Alignment(horizontal='center', vertical='center')
        merged_cellb2.font = Font( size=14, bold=False)

        # Apply the thick bottom border to each cell in the range B2:O2
        for col in range(2, 16):  # Columns B (2) to O (15)
            cell = Scenarios.cell(row=2, column=col)
            cell.border = thick_bottom_border  # Apply the thick border
        
        #Set Scenario Title
        Scenarios['B10'] = 'ECONOMIC SCENARIOS'
        b10 = Scenarios['B10']
        b10.font = Font( size=12, bold=False)
        #Set Scenario SubTitle
        Scenarios['B12'] = 'Interest Rates'
        # Write cases
        Scenarios['C14'] = '=Assumptions!C11'
        Scenarios['C15'] = '=Assumptions!C12'
        Scenarios['C16'] = '=Assumptions!C13'
        Scenarios['B14'] = '=Assumptions!G11'
        Scenarios['B16'] = '=Assumptions!G13'
        
        #Set Scenario SubTitle
        Scenarios['B18'] = 'GDP Growth'
        #Write Cases
        Scenarios['C20'] = '=Assumptions!C17'
        Scenarios['C21'] = '=Assumptions!C18'
        Scenarios['C22'] = '=Assumptions!C19'
        Scenarios['B20'] = '=Assumptions!G17'
        Scenarios['B22'] = '=Assumptions!G19'
        
        # Set Political switch
        Scenarios['B26'] = 'POLITICAL SWITCH:'
        Scenarios['D26'] = '1'
        
        #Set Political Scenario Title
        Scenarios['B29'] = 'POLITICS SCENARIOS'
        b29 = Scenarios['B29']
        b29.font = Font( size=12, bold=False)
        #Set Political Scenario SubTitle
        Scenarios['B31'] = 'Tax Rate'
        # Write cases
        Scenarios['C33'] = '=Assumptions!J11'
        Scenarios['C34'] = '=Assumptions!J12'
        Scenarios['C35'] = '=Assumptions!J13'
        Scenarios['B33'] = '=Assumptions!M11'
        Scenarios['B34'] = '=Assumptions!M12'
        Scenarios['B35'] = '=Assumptions!M13'
        
        Scenarios['G33'] = '=Assumptions!M11'
        Scenarios['G34'] = '=Assumptions!M12'
        Scenarios['G35'] = '=Assumptions!M13'
        
        # Set Sales switch
        Scenarios['B39'] = 'SALES SWITCH:'
        Scenarios['D39'] = '1'
        #Set Sales Scenario SubTitle
        Scenarios['B42'] = 'SALES SCENARIOS'
        b42 = Scenarios['B42']
        b42.font = Font( size=12, bold=False)
        Scenarios['B44'] = 'Segment A Growth'
        # Write cases
        Scenarios['C46'] = '=Assumptions!J17'
        Scenarios['C47'] = '=Assumptions!J18'
        Scenarios['C48'] = '=Assumptions!J19'
        Scenarios['B46'] = '=Assumptions!M17'
        Scenarios['B48'] = '=Assumptions!M19'
        
        Scenarios['B50'] = 'Segment B Growth'
        # Write cases
        Scenarios['C52'] = '=Assumptions!C23'
        Scenarios['C53'] = '=Assumptions!C24'
        Scenarios['C54'] = '=Assumptions!C25'
        Scenarios['B52'] = '=Assumptions!G23'
        Scenarios['B54'] = '=Assumptions!G25'
        
        #Apply live interest rate case formulas
        Scenarios['G12'] = '=CHOOSE($D$6,G14,G15,G16)'
        Scenarios['H12'] = '=CHOOSE($D$6,H14,H15,H16)'
        Scenarios['I12'] = '=CHOOSE($D$6,I14,I15,I16)'
        Scenarios['J12'] = '=CHOOSE($D$6,J14,J15,J16)'
        Scenarios['K12'] = '=CHOOSE($D$6,K14,K15,K16)'
        
        #Apply live inflation case formulas
        Scenarios['G18'] = '=CHOOSE($D$6,G20,G21,G22)'
        Scenarios['H18'] = '=CHOOSE($D$6,H20,H21,H22)'
        Scenarios['I18'] = '=CHOOSE($D$6,I20,I21,I22)'
        Scenarios['J18'] = '=CHOOSE($D$6,J20,J21,J22)'
        Scenarios['K18'] = '=CHOOSE($D$6,K20,K21,K22)'
        
        #Apply live tax rate case formulas
        Scenarios['G31'] = '=CHOOSE($D$26,G33,G34,G35)'
        
        #Apply live Segment A case formulas
        Scenarios['G44'] = '=CHOOSE($D$39,G46,G47,G48)'
        Scenarios['H44'] = '=CHOOSE($D$39,H46,H47,H48)'
        Scenarios['I44'] = '=CHOOSE($D$39,I46,I47,I48)'
        Scenarios['J44'] = '=CHOOSE($D$39,J46,J47,J48)'
        Scenarios['K44'] = '=CHOOSE($D$39,K46,K47,K48)'
        
        #Apply live Segment B case formulas
        Scenarios['G50'] = '=CHOOSE($D$39,G52,G53,G54)'
        Scenarios['H50'] = '=CHOOSE($D$39,H52,H53,H54)'
        Scenarios['I50'] = '=CHOOSE($D$39,I52,I53,I54)'
        Scenarios['J50'] = '=CHOOSE($D$39,J52,J53,J54)'
        Scenarios['K50'] = '=CHOOSE($D$39,K52,K53,K54)'
        
        # Define a dotted border
        apply_dotted_border(Scenarios, 'G14', 'K16')
        apply_dotted_border(Scenarios, 'G20', 'K22')
        apply_dotted_border(Scenarios, 'G33', 'G35')
        apply_dotted_border(Scenarios, 'G46', 'K48')
        apply_dotted_border(Scenarios, 'G52', 'K54')
        apply_dotted_border(Model, 'B168', 'O168')
        # Define a thin border
        apply_thin_border(Scenarios, 'B5', 'E7')
        apply_thin_border(Scenarios, 'G12', 'K12')
        apply_thin_border(Scenarios, 'G18', 'K18')
        apply_thin_border(Scenarios, 'B25', 'E27')
        apply_thin_border(Scenarios, 'G31', 'G31')
        apply_thin_border(Scenarios, 'B38', 'E40')
        apply_thin_border(Scenarios, 'G44', 'K44')
        apply_thin_border(Scenarios, 'G50', 'K50')
        apply_thin_border(Model, 'B52', 'O55')
        apply_thin_border(Model, 'B125', 'O127')
        apply_thin_border(Model, 'B135', 'F136')
        apply_thin_border(Model, 'E140', 'E144')
        apply_thin_border(Model, 'C153', 'E153')
        apply_thin_border(Model, 'C275', 'E276')
        apply_thin_border(Model, 'C285', 'E301')
        # Define boxes for assumptions
        apply_thin_border(Assumptions, 'C4', 'H7')
        apply_thin_border(Assumptions, 'J4', 'N7')
        apply_thin_border(Assumptions, 'C9', 'H13')
        apply_thin_border(Assumptions, 'J9', 'N13')
        apply_thin_border(Assumptions, 'C15', 'H19')
        apply_thin_border(Assumptions, 'J15', 'N19')
        apply_thin_border(Assumptions, 'C21', 'H25')
        apply_thin_border(Assumptions, 'J21', 'N25')
        apply_thin_border(Assumptions, 'C27', 'N59')
        # Define page breaks
        thin_line(Scenarios, 'B23', 'O23')
        thin_line(Scenarios, 'B36', 'O36')
        thin_line(Model, 'B10', 'O10')
        thin_line(Model, 'B15', 'O15')
        
        # Define page end
        thick_line(Scenarios, 'G5', 'K5')
        thick_line(Scenarios, 'B55', 'O55')
        thick_line(Model, 'K5', 'O5')
        thick_line(Model, 'B21', 'O21')
        thick_line(Model, 'K27', 'O27')
        thick_line(Model, 'H31', 'O31')
        thick_line(Model, 'H35', 'O35')
        thick_line(Model, 'H40', 'O40')
        thick_line(Model, 'H43', 'O43')
        thick_line(Model, 'H46', 'O46')
        thick_line(Model, 'H49', 'O49')
        thick_line(Model, 'B56', 'O56')
        thick_line(Model, 'K62', 'O62')
        thick_line(Model, 'H68', 'O68')
        thick_line(Model, 'H72', 'O72')
        thick_line(Model, 'H74', 'O74')
        thick_line(Model, 'H80', 'O80')
        thick_line(Model, 'H84', 'O84')
        thick_line(Model, 'H86', 'O86')
        thick_line(Model, 'H93', 'O93')
        thick_line(Model, 'H95', 'O95')
        thick_line(Model, 'B98', 'O98')
        thick_line(Model, 'K103', 'O103')
        thick_line(Model, 'H109', 'O109')
        thick_line(Model, 'H114', 'O114')
        thick_line(Model, 'H122', 'O122')
        thick_line(Model, 'B128', 'O128')
        thick_line(Model, 'K134', 'O134')
        thick_line(Model, 'H144', 'O144')
        thick_line(Model, 'B146', 'O146')
        thick_line(Model, 'K152', 'O152')
        thick_line(Model, 'B159', 'O159')
        thick_line(Model, 'K165', 'O165')
        thick_line(Model, 'I179', 'O179')
        thick_line(Model, 'J181', 'O181')
        thick_line(Model, 'B183', 'O183')
        thick_line(Model, 'K189', 'O189')
        thick_line(Model, 'J195', 'O195')
        thick_line(Model, 'K198', 'O198')
        thick_line(Model, 'J203', 'O203')
        thick_line(Model, 'K206', 'O206')
        thick_line(Model, 'K214', 'O214')
        thick_line(Model, 'J218', 'O218')
        thick_line(Model, 'J237', 'O237')
        thick_line(Model, 'K231', 'O231')
        thick_line(Model, 'K241', 'O241')
        thick_line(Model, 'J246', 'O246')
        thick_line(Model, 'K250', 'O250')
        thick_line(Model, 'B269', 'O269')
        thick_line(Model, 'J255', 'O255')
        thick_line(Model, 'K223', 'O223')
        thick_line(Model, 'J260', 'O260')
        thick_line(Model, 'K221', 'O221')
        thick_line(Model, 'J267', 'O267')
        thick_line(Model, 'K275', 'O275')
        thick_line(Model, 'K282', 'O282')

    print(f'Data has been written to {output_path}')


# In[14]:


# Example usage
input_path = r"C:\Users\thoma\Desktop\Projects\Project Oversight\Data Annual\Combined_Annual_Data.xlsx"
output_path = 'DCF_Framework.xlsx'
sheet_name = 'NVDA'
company_name = 'NVIDIA Corporation'
years = [2025, 2026, 2027, 2028, 2029]

create_dcf_file(input_path, output_path, sheet_name, company_name, years)


# In[ ]:





# In[104]:





# In[121]:





# In[97]:





# In[ ]:




